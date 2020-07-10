import sys, os, time
import urllib.request
import getpass
import py2exe
import requests
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import Select

path = input("Please enter file path: ")
wk = openpyxl.load_workbook(path)
sh = wk['Sheet1']
rows = sh.max_row
User_email = input("Your Email: ")
password = getpass.getpass("Your Password: ")
seller_name = input("Attribute Group: ")
#User_password = input("Your Password: ")
CSV_path1 = input("Please enter location to save file: ")
CSV_path =  CSV_path1.replace('/','//')
if getattr(sys, 'frozen', False):
    chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")
    driver = webdriver.Chrome(chromedriver_path)
else:
    driver = webdriver.Chrome()
driver.get('https://backend.goto.com.pk/catalog/products/index')
time.sleep(5)
Email = driver.find_element_by_id('usersform-email').send_keys(User_email)
Password = driver.find_element_by_id('usersform-password').send_keys(password)
Enter = driver.find_element_by_id('usersform-password').send_keys( u'\ue007')
time.sleep(2)
data = []
for i in range (1, rows+1):
         val = sh.cell(i,1).value
         print(val)
         col1 = sh.max_column
         print(col1)
         catalog_search = driver.find_element_by_xpath('/html/body/div[1]/aside/div/section/ul/li[4]/a').click()
         time.sleep(2)
         catalog_search12 = driver.find_element_by_xpath('/html/body/div[1]/aside/div/section/ul/li[4]/ul/li[6]/a').click()
         time.sleep(2)
         catalog_search136 = driver.find_element_by_xpath('/html/body/div[1]/div/section[1]/div/a[2]').click()
         time.sleep(2)
         Clear_value = driver.find_element_by_xpath('//*[@id="searchModal"]/div/div/div[2]/div/div[3]/div/input').clear()
         Enter_SKU = driver.find_element_by_xpath('//*[@id="searchModal"]/div/div/div[2]/div/div[3]/div/input').send_keys(val)
         Enter_SKU_Search = driver.find_element_by_xpath('//*[@id="searchModal"]/div/div/div[2]/div/div[3]/div/input').send_keys(u'\ue007')
         time.sleep(5)
         edit_click = driver.find_element_by_xpath('//*[@id="w0"]/div/div[2]/table/tbody/tr/td[9]/a[@title="Update Product"]').click()
         time.sleep(2)
         dropdown = driver.find_element_by_xpath('//*[@id="products-attribute_group_id"]').click()
         time.sleep(3)
         dropdown_search = driver.find_element_by_xpath('//*[@id="products-attribute_group_id"]').send_keys(seller_name)
         time.sleep(2)
         dropdown_search1 = driver.find_element_by_xpath('//*[@id="products-attribute_group_id"]').click()
         time.sleep(2)
         dup_submit = driver.find_element_by_xpath('//*[@id="sub"]').click()
         time.sleep(5)
         data.append((val))
         import pandas as pd
         df = pd.DataFrame(data,columns =['Done SKU'])
         df.to_csv(CSV_path+'//DATA.csv',index=False,encoding='utf-8')
         time.sleep(5)
         driver.get('https://backend.goto.com.pk/dashboards/enlisting')
##driver.close()
##username_drop = driver.find_element_by_xpath('/html/body/div[1]/header/nav/div/ul/li[@class="dropdown user user-menu"]/a').click()
##userlogout = driver.find_element_by_xpath('/html/body/div[1]/header/nav/div/ul/li/ul/li[3]/a').click()         
##         try:
##             size39 = drp.select_by_visible_text('Beauty Products').text
##         except:
##             try:
##                 Size_39_stock = driver.find_element_by_xpath('//*[@id="products-attribute_group_id"]/option[12]').text
##             except:
##                 Size_39_stock = '39 Size unavailable'
##                 pass
##         time.sleep(2)       
##        
##         linked_prod = driver.find_element_by_xpath('//*[@id="w0"]/div[1]/div/div/ul/li[9]/a[contains(.,"Linked")]').click()
##         k=0
##         for j in range (2, col1+1):             
##             val1 = sh.cell(1,j).value
##             print(str(j)+","+str(k))
##             custom_add = driver.find_element_by_xpath('//*[@id="w1"]/table/tbody/tr/td[@class="list-cell__button"]').click()
##             time.sleep(2)
##             custom_input = driver.find_element_by_xpath('//*[@id="w1"]/table/tbody/tr/td[1]/div/input[@id="dynamicmodel-childskus-%d"]' % (k,)).send_keys(val1)
##             time.sleep(2)
##             k=k+1
##         field_cut = driver.find_element_by_xpath('//*[@id="w1"]/table/tbody/tr['+col1+']/td[2]').click()
##         time.sleep(1)
##         custom_submit = driver.find_element_by_xpath('//*[@id="w0"]/div[2]/button').click()
##         time.sleep(2)    
##         a=1
##         while(a==1):
##             try:
##                driver.find_element_by_xpath('/html/body/div/div/section[2]/div[contains(@class,"alert-success")]')
##                a=2
##                time.sleep(3)
##                product_section = driver.find_element_by_xpath('/html/body/div/aside/section/ul/li[4]/ul/li[@class="active"]/a').click()
##             except:
##                 pass
##//*[@id="product-specs"]/div[3]/ul/li/ul/li/em[contains(@class,'product-specs__highlights-sub-title') and contains(text(),'Screen Size')]//following::span[1]
