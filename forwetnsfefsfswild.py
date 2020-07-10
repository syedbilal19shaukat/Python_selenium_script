import urllib.request
import openpyxl
import time
import times
from random import *
import random,string
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from boltons import iterutils
import itertools    
wk = openpyxl.load_workbook(r'C:\Users\Bilal Shaukat\Desktop\beauti-indution.xlsx')
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column
data = []
driver = webdriver.Chrome()
driver.maximize_window()
#browser = webdriver.Chrome()
for i in range (1,  rows+1):
         val = sh.cell(i,1).value
         time.sleep(5)
         driver.get(val)
         j=1
         while driver.find_elements_by_xpath('//div[@class="swatch-attribute-options clearfix"]/div'):
             try:
                Image7 = driver.find_element_by_xpath('//div[@class="fotorama__nav__shaft"]/div[2]/div/img').get_attribute('src')
                print(Image7)
                len7 = 10
                letters7 = string.ascii_lowercase
                Image7=Image7.replace("jpg_80x80q80.jpg_.webp","jpg")
                img7_name = ''.join(random.choice(letters7) for i in range(len7))
                full7_name = str(img7_name) + '.jpg'
                file7_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full7_name
                urllib.request.urlretrieve(Image7,file7_path)
             except:
                Image7 = 'No Image'
                full7_name = 'No Image'
                print(Image7)
                pass
             try:
                Image8 = driver.find_element_by_xpath('//div[@class="fotorama__nav__shaft"]/div[3]/div/img').get_attribute('src')
                print(Image8)
                len8 = 10
                letters8 = string.ascii_lowercase
                Image8=Image8.replace("jpg_80x80q80.jpg_.webp","jpg")
                img8_name = ''.join(random.choice(letters8) for i in range(len8))
                full8_name = str(img8_name) + '.jpg'
                file8_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full8_name
                urllib.request.urlretrieve(Image8,file8_path)
             except:
                Image8 = 'No Image'
                full8_name = 'No Image'
                print(Image8)
                pass

                
         data.append((full7_name,full8_name,val))
         import pandas as pd
         df = pd.DataFrame(data,columns =['Image7','Image8','PageURL'])
         df.to_csv('E:\\Crawling-py-Data\\extractor\\New folder\\file.csv',index=False,encoding='utf-8')
         
           
        
            
          
           
            
            
            
            

            
                        
         
        

   



