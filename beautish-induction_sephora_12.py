import urllib.request
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
wk = openpyxl.load_workbook(r'C:\Users\gotoadmin\Desktop\34074.xlsx')
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column
data = []
driver = webdriver.Chrome()
driver.maximize_window()
#browser = webdriver.Chrome()
for i in range (1, rows+1):
         val = sh.cell(i,1).value
         driver.get(val)
         time.sleep(4)
         try:
             Image1 = driver.find_element_by_xpath('//div[@class="row transitionfx"]/div[1]/div/img[contains(@class,"img-responsive")]').get_attribute('src')
             print(Image1)
##            len1 = 10
##            letters1 = string.ascii_lowercase
##            ##Image1=Image1.replace("a877073a111072cd12066c1bf9b87579","926507dc7f93631a094422215b778fe0")
##            img1_name = ''.join(random.choice(letters1) for i in range(len1))
##            full1_name = str(img1_name) + '.jpg'
##            file1_path = 'D:\\crawling\\New folder_1\\' + full1_name
##            urllib.request.urlretrieve(Image1,file1_path)
         except:
##            Image1 = 'No Image'
##            full1_name = 'No Image'
##            print(Image1)
            pass
##         try:
##            time.sleep(8)
##            Image2 = driver.find_element_by_xpath('//*[@id="maincontent"]/div[2]/div/div[1]/div[1]/div[2]/div[2]/div[2]/div[2]/div/div[2]/div[3]/div/img').get_attribute('src')
##            print(Image2)
##            len2 = 10
##            letters2 = string.ascii_lowercase
##            Image2=Image2.replace("a877073a111072cd12066c1bf9b87579","926507dc7f93631a094422215b778fe0")
##            img2_name = ''.join(random.choice(letters2) for i in range(len2))
##            full2_name = str(img2_name) + '.jpg'
##            file2_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full2_name
##            urllib.request.urlretrieve(Image2,file2_path)
##         except:
##            Image2 = 'No Image'
##            full2_name = 'No Image'
##            print(Image2)
##            pass
##         try:
##            time.sleep(8)
##            Image3 = driver.find_element_by_xpath('//*[@id="maincontent"]/div[2]/div/div[1]/div[1]/div[2]/div[2]/div[2]/div[2]/div/div[2]/div[4]/div/img').get_attribute('src')
##            print(Image3)
##            len3 = 10
##            letters3 = string.ascii_lowercase
##            Image3=Image3.replace("a877073a111072cd12066c1bf9b87579","926507dc7f93631a094422215b778fe0")
##            img3_name = ''.join(random.choice(letters3) for i in range(len3))
##            full3_name = str(img3_name) + '.jpg'
##            file3_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full3_name
##            urllib.request.urlretrieve(Image3,file3_path)
##         except:
##            Image3 = 'No Image'
##            full3_name = 'No Image'
##            print(Image3)
##            pass
##         try:
##            time.sleep(8)
##            Image4 = driver.find_element_by_xpath('//ul[@class="buttonspan"]/li[4]/a/img').get_attribute('src')
##            print(Image4)
##            len4 = 10
##            letters4 = string.ascii_lowercase
##            Image4=Image4.replace("/120","/950")
##            img4_name = ''.join(random.choice(letters4) for i in range(len4))
##            full4_name = str(img4_name) + '.jpg'
##            file4_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full4_name
##            urllib.request.urlretrieve(Image4,file4_path)
##         except:
##            Image4 = 'No Image'
##            full4_name = 'No Image'
##            print(Image4)
##            pass
##         try:
##            Image5 = driver.find_element_by_xpath('//div[@class="next-slick-list"]/div/div[5]/div/img').get_attribute('src')
##            print(Image5)
##            len5 = 10
##            letters5 = string.ascii_lowercase
##            Image5=Image5.replace("jpg_80x80q80.jpg_.webp","jpg")
##            img5_name = ''.join(random.choice(letters5) for i in range(len5))
##            full5_name = str(img5_name) + '.jpg'
##            file5_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full5_name
##            urllib.request.urlretrieve(Image5,file5_path)
##         except:
##            Image5 = 'No Image'
##            full5_name = 'No Image'
##            print(Image5)
##            pass
##         try:
##            Image6 = driver.find_element_by_xpath('//div[@class="next-slick-list"]/div/div[6]/div/img').get_attribute('src')
##            print(Image6)
##            len6 = 10
##            letters6 = string.ascii_lowercase
##            Image6=Image6.replace("jpg_80x80q80.jpg_.webp","jpg")
##            img6_name = ''.join(random.choice(letters6) for i in range(len6))
##            full6_name = str(img6_name) + '.jpg'
##            file6_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full6_name
##            urllib.request.urlretrieve(Image6,file6_path)
##         except:
##            Image6 = 'No Image'
##            full6_name = 'No Image'
##            print(Image6)
##            pass
##         try:
##            Image7 = driver.find_element_by_xpath('//div[@class="next-slick-list"]/div/div[7]/div/img').get_attribute('src')
##            print(Image7)
##            len7 = 10
##            letters7 = string.ascii_lowercase
##            Image7=Image7.replace("jpg_80x80q80.jpg_.webp","jpg")
##            img7_name = ''.join(random.choice(letters7) for i in range(len7))
##            full7_name = str(img7_name) + '.jpg'
##            file7_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full7_name
##            urllib.request.urlretrieve(Image7,file7_path)
##         except:
##            Image7 = 'No Image'
##            full7_name = 'No Image'
##            print(Image7)
##            pass
##         try:
##            Image8 = driver.find_element_by_xpath('//div[@class="next-slick-list"]/div/div[8]/div/img').get_attribute('src')
##            print(Image8)
##            len8 = 10
##            letters8 = string.ascii_lowercase
##            Image8=Image8.replace("jpg_80x80q80.jpg_.webp","jpg")
##            img8_name = ''.join(random.choice(letters8) for i in range(len8))
##            full8_name = str(img8_name) + '.jpg'
##            file8_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full8_name
##            urllib.request.urlretrieve(Image8,file8_path)
##         except:
##            Image8 = 'No Image'
##            full8_name = 'No Image'
##            print(Image8)
##            pass

         data.append((Image1,val))
         import pandas as pd
         df = pd.DataFrame(data,columns =['Image1','PageURL'])
         df.to_csv('D:\\crawling\\New folder_1\\skinlee.csv',index=False,encoding='utf-8')
         
           
        
            
          
           
            
            
            
            

            
                        
         
        

   



