#----------- daraz crawl
import sys, os, time
import urllib.request
import getpass
import requests
import openpyxl
import sys, os, time
import re
import requests
from array import array
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

path = input("Please enter file path: ")
wk = openpyxl.load_workbook(path)
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column
filename=input("Please Enter file name: ")
CSV_path1 = input("Please enter location to save file: ")
CSV_path =  CSV_path1.replace('/','//')
if getattr(sys, 'frozen', False):
    chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")
    driver = webdriver.Chrome(chromedriver_path)
else:
    driver = webdriver.Chrome()
driver.maximize_window
wb=openpyxl.Workbook()
sheet=wb.active
sheet.cell(row=1, column=1).value="Links"
sheet.cell(row=1, column=2).value="name"
sheet.cell(row=1, column=3).value="brand"
sheet.cell(row=1, column=4).value="primary_category"
sheet.cell(row=1, column=5).value="promotion_start_date"
sheet.cell(row=1, column=6).value="promotion_price"
sheet.cell(row=1, column=7).value="promotion_msrp"
sheet.cell(row=1, column=8).value="promotion_quantity"
sheet.cell(row=1, column=9).value="promotion_end_date"
sheet.cell(row=1, column=10).value="promotion_payment_method"
sheet.cell(row=1, column=11).value="promotion_uses_per_customer"
sheet.cell(row=1, column=12).value="promotion_type"
sheet.cell(row=1, column=13).value="promotion_bins"
sheet.cell(row=1, column=14).value="variation"
sheet.cell(row=1, column=15).value="parent_sku"
sheet.cell(row=1, column=16).value="vendor_sku"
sheet.cell(row=1, column=17).value="quantity"
sheet.cell(row=1, column=18).value="warehouse_quantity"
sheet.cell(row=1, column=19).value="long_description"
sheet.cell(row=1, column=20).value="short_description"
sheet.cell(row=1, column=21).value="specifications"
sheet.cell(row=1, column=22).value="refund_policy"
sheet.cell(row=1, column=23).value="product_warranty_short"
sheet.cell(row=1, column=24).value="product_warranty_long"
sheet.cell(row=1, column=25).value="price"
sheet.cell(row=1, column=26).value="code"
sheet.cell(row=1, column=27).value="main_image"
sheet.cell(row=1, column=28).value="image_2"
sheet.cell(row=1, column=29).value="image_3"
sheet.cell(row=1, column=30).value="image_4"
sheet.cell(row=1, column=31).value="image_5"
sheet.cell(row=1, column=32).value="image_6"
sheet.cell(row=1, column=33).value="image_7"
sheet.cell(row=1, column=34).value="image_8"
sheet.cell(row=1, column=35).value="meta_title"
sheet.cell(row=1, column=36).value="meta_keywords"
sheet.cell(row=1, column=37).value="meta_description"
sheet.cell(row=1, column=38).value="msrp"
sheet.cell(row=1, column=39).value="search_weight"
sheet.cell(row=1, column=40).value="shipping_cost_class"
sheet.cell(row=1, column=41).value="free_shipping"
sheet.cell(row=1, column=42).value="color"
sheet.cell(row=1, column=43).value="product_weight"
sheet.cell(row=1, column=44).value="color_family"
sheet.cell(row=1, column=45).value="product_keywords"
sheet.cell(row=1, column=46).value="delivery_time"
sheet.cell(row=1, column=47).value="youtube_id"
sheet.cell(row=1, column=48).value="shipping_height"
sheet.cell(row=1, column=49).value="shipping_length"
sheet.cell(row=1, column=50).value="shipping_width"
sheet.cell(row=1, column=51).value="shipping_weight"
sheet.cell(row=1, column=52).value="sku_city"
sheet.cell(row=1, column=53).value="upc"
sheet.cell(row=1, column=54).value="ean"
sheet.cell(row=1, column=55).value="what_is_in_the_box"
sheet.cell(row=1, column=56).value="breadcrumb"
sheet.cell(row=1, column=57).value="ampere"
sheet.cell(row=1, column=58).value="audio_format"
sheet.cell(row=1, column=59).value="automatic_redial_status"
sheet.cell(row=1, column=60).value="back_camera_megapixel"
sheet.cell(row=1, column=61).value="back_camera_flash_status"
sheet.cell(row=1, column=62).value="back_camera_sensor_status"
sheet.cell(row=1, column=63).value="back_camera_status"
sheet.cell(row=1, column=64).value="battery_cells_no"
sheet.cell(row=1, column=65).value="battery_chargeable_status"
sheet.cell(row=1, column=66).value="battery_charging_time"
sheet.cell(row=1, column=67).value="battery_included"
sheet.cell(row=1, column=68).value="battery_input_current"
sheet.cell(row=1, column=69).value="battery_input_voltage"
sheet.cell(row=1, column=70).value="battery_maximum_talk_time"
sheet.cell(row=1, column=71).value="battery_ouput_current"
sheet.cell(row=1, column=72).value="battery_output_voltage"
sheet.cell(row=1, column=73).value="battery_rated_input_power"
sheet.cell(row=1, column=74).value="battery_rated_output_power"
sheet.cell(row=1, column=75).value="battery_standby_time"
sheet.cell(row=1, column=76).value="builtin_camera_status"
sheet.cell(row=1, column=77).value="cellular_generation"
sheet.cell(row=1, column=78).value="cellular_connectivity_tech"
sheet.cell(row=1, column=79).value="memory_card_class"
sheet.cell(row=1, column=80).value="conference_call_status"
sheet.cell(row=1, column=81).value="connectivity"
sheet.cell(row=1, column=82).value="connectivity_type"
sheet.cell(row=1, column=83).value="control_and_adjustments_volume"
sheet.cell(row=1, column=84).value="deep_bass_status"
sheet.cell(row=1, column=85).value="driver_material_type"
sheet.cell(row=1, column=86).value="expansion_handsets_status"
sheet.cell(row=1, column=87).value="extendable_memory"
sheet.cell(row=1, column=88).value="extended_physical_height"
sheet.cell(row=1, column=89).value="extended_physical_length"
sheet.cell(row=1, column=90).value="extended_physical_width"
sheet.cell(row=1, column=91).value="external_storage"
sheet.cell(row=1, column=92).value="fast_charging_status"
sheet.cell(row=1, column=93).value="frequency"
sheet.cell(row=1, column=94).value="front_camera_megapixel"
sheet.cell(row=1, column=95).value="front_camera_flash"
sheet.cell(row=1, column=96).value="front_camera_sensor"
sheet.cell(row=1, column=97).value="front_camera_status"
sheet.cell(row=1, column=98).value="gps_status"
sheet.cell(row=1, column=99).value="headphone_type"
sheet.cell(row=1, column=100).value="input_output_connector_type"
sheet.cell(row=1, column=101).value="impedance"
sheet.cell(row=1, column=102).value="input_voltage"
sheet.cell(row=1, column=103).value="keypad_type"
sheet.cell(row=1, column=104).value="loud_speaker_status"
sheet.cell(row=1, column=105).value="mic_status"
sheet.cell(row=1, column=106).value="nfc_status"
sheet.cell(row=1, column=107).value="no_of_telephone_lines"
sheet.cell(row=1, column=108).value="notification"
sheet.cell(row=1, column=109).value="operating_system"
sheet.cell(row=1, column=110).value="os_version"
sheet.cell(row=1, column=111).value="physical_height"
sheet.cell(row=1, column=112).value="physical_length"
sheet.cell(row=1, column=113).value="physical_width"
sheet.cell(row=1, column=114).value="secondary_sim_slot_status"
sheet.cell(row=1, column=115).value="sensors"
sheet.cell(row=1, column=116).value="sim_slot"
sheet.cell(row=1, column=117).value="simcard_quantity"
sheet.cell(row=1, column=118).value="sound_type"
sheet.cell(row=1, column=119).value="webcam_status"
sheet.cell(row=1, column=120).value="battery_size"
sheet.cell(row=1, column=121).value="bluetooth_version"
sheet.cell(row=1, column=122).value="cable_type"
sheet.cell(row=1, column=123).value="compatible_device"
sheet.cell(row=1, column=124).value="processor_model"
sheet.cell(row=1, column=125).value="form_factor"
sheet.cell(row=1, column=126).value="gpu_model"
sheet.cell(row=1, column=127).value="gpu_speed"
sheet.cell(row=1, column=128).value="internal_storage"
sheet.cell(row=1, column=129).value="keyboard_localization"
sheet.cell(row=1, column=130).value="material_family"
sheet.cell(row=1, column=131).value="model"
sheet.cell(row=1, column=132).value="series_no"
sheet.cell(row=1, column=133).value="processor_speed"
sheet.cell(row=1, column=134).value="processor_type"
sheet.cell(row=1, column=135).value="product_material"
sheet.cell(row=1, column=136).value="battery_removeable_status"
sheet.cell(row=1, column=137).value="run_time"
sheet.cell(row=1, column=138).value="screen_size"
sheet.cell(row=1, column=139).value="screen_type"
sheet.cell(row=1, column=140).value="touch_screen_status"
sheet.cell(row=1, column=141).value="product_type"
sheet.cell(row=1, column=142).value="noise_cancellation"
sheet.cell(row=1, column=143).value="video_format_support"
sheet.cell(row=1, column=144).value="battery_type"
sheet.cell(row=1, column=145).value="ram_rated_memory_speed"
sheet.cell(row=1, column=146).value="ram_rated_memory"
sheet.cell(row=1, column=147).value="bluetooth"
x=0
urls=[]
browser = webdriver.Chrome()
for i in range(len(urls)):
    print("Link: "+str(i+1))
    sheet.cell(row=x+2, column=1).value=urls[i]
    try:
        res = requests.get(urls[i])
    except:
        pass
    time.sleep(5)
    soup = BeautifulSoup(res.content,'html.parser')
    try:
        sheet.cell(row=x+2, column=57).value=soup.find("div" , {"id" : "specs-list"}).text
        
    except:
        pass
##    try:
##        sheet.cell(row=x+2, column=60).value=driver.find_element_by_xpath('//table[7]/tbody/tr/td[contains(@data-spec,"cam1modules")][1]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=66).value=driver.find_element_by_xpath('//table[12]/tbody/tr[2]/td[2]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=70).value=driver.find_element_by_xpath('//table[12]/tbody/tr[3]/td[contains(@data-spec,"battalktime1")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=77).value=driver.find_element_by_xpath('//table[1]/tbody/tr/td/a[contains(text(),"2G bands")] | //table[1]/tbody/tr/td/a[contains(text(),"3G bands")] | //table[1]/tbody/tr/td/a[contains(text(),"4G bands")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=78).value=driver.find_element_by_xpath('//table[1]/tbody/tr/td/a[contains(text(),"GPRS")] | //table[1]/tbody/tr/td/a[contains(text(),"EDGE")] | //table[1]/tbody/tr[3]/td[2]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=82).value=driver.find_element_by_xpath('//table[10]/tbody/tr[2]/td').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=91).value=driver.find_element_by_xpath('//table[6]/tbody/tr[1]/td[contains(@data-spec,"memoryslot")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=94).value=driver.find_element_by_xpath('//table[8]/tbody/tr[1]/td[contains(@data-spec,"cam2modules")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=109).value=driver.find_element_by_xpath('//table[5]/tbody/tr[1]/td[contains(@data-spec,"os")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=110).value=driver.find_element_by_xpath('//table[5]/tbody/tr[1]/td[contains(@data-spec,"os")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=111).value=driver.find_element_by_xpath('//table[3]/tbody/tr[1]/td[contains(@data-spec,"dimensions")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=112).value=driver.find_element_by_xpath('//table[3]/tbody/tr[1]/td[contains(@data-spec,"dimensions")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=113).value=driver.find_element_by_xpath('//table[3]/tbody/tr[1]/td[contains(@data-spec,"dimensions")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=115).value=driver.find_element_by_xpath('//table[11]/tbody/tr/td[contains(@data-spec,"sensors")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=116).value=driver.find_element_by_xpath('//table[3]/tbody/tr/td[contains(@data-spec,"sim")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=121).value=driver.find_element_by_xpath('//table[10]/tbody/tr[2]/td[contains(@data-spec,"bluetooth")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=126).value=driver.find_element_by_xpath('//table[5]/tbody/tr[4]/td[contains(@data-spec,"gpu")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=128).value=driver.find_element_by_xpath('//table[6]/tbody/tr[2]/td[contains(@data-spec,"internalmemory")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=133).value=driver.find_element_by_xpath('//table[5]/tbody/tr[3]/td[contains(@data-spec,"cpu")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=134).value=driver.find_element_by_xpath('//table[5]/tbody/tr[3]/td[contains(@data-spec,"cpu")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=138).value=driver.find_element_by_xpath('//table[4]/tbody/tr/td[contains(@data-spec,"displaysize")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=139).value=driver.find_element_by_xpath('//table[4]/tbody/tr/td[contains(@data-spec,"displaytype")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=144).value=driver.find_element_by_xpath('//table[12]/tbody/tr/td[contains(@data-spec,"batdescription1")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=145).value=driver.find_element_by_xpath('//table[6]/tbody/tr[2]/td[contains(@data-spec,"internalmemory")]').text
##    except:
##        pass
##    try:
##        sheet.cell(row=x+2, column=146).value=driver.find_element_by_xpath('//table[6]/tbody/tr[2]/td[contains(@data-spec,"internalmemory")]').text
##    except:
##        pass

    if len(sizearray)!=0:
        x=z
    else:
        x=x+1
    wb.save(filename+'.xlsx')
wb.save(filename+'.xlsx')
print("Download Compeltes...")
