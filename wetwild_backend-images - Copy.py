import sys, os, time
import urllib.request
import getpass
import py2exe
import requests
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from boltons import iterutils
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
wk = openpyxl.load_workbook(r'C:\Users\Bilal Shaukat\Desktop\beauti-indution.xlsx')
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column
data = []
driver = webdriver.Chrome()
driver.maximize_window()
time.sleep(2)
for i in range (1, rows+1):
         val = sh.cell(i,1).value
         driver.get(val)
         time.sleep(15)
         try:
             cross = driver.find_element_by_xpath('//html[@class="no-touch"]/body//div[@class="campaign "]/div/div[contains(@class, "popup-close is-solid js-offer-close")]')
             cross.click()
         except:
             pass
         time.sleep(5)
         j=1
         imagename=[]
         skuu=[]
         for link in driver.find_elements_by_xpath('//div[contains(@class,"swatch-attribute-options")]/div'): 
             link = driver.find_element_by_xpath('//div[contains(@class,"swatch-attribute-options")]/div[%d]' % (j,)).click()
             time.sleep(10)
             Image = dict()
             sampleImage = dict()
             fullname = dict()
             filepath = dict()
             k=1
             for images in driver.find_elements_by_xpath('//div[@class="fotorama__nav fotorama__nav--thumbs"]/div/div[contains(@class,"fotorama__nav__frame fotorama__nav__frame--thumb")]'):
                 sampleImage[k] = driver.find_element_by_xpath('//div[@class="fotorama__nav__shaft"]/div[contains(@data-gallery-role,"nav-frame")][%d]/div/img' % (k,)).get_attribute('src')
                 Image[k] = sampleImage[k].replace('f9c7fbe9b524c081a3ccf800cbd963eb','c687aa7517cf01e65c009f6943c2b1e9')
                 print(Image[k])
                 characters = 10
                 letters = string.ascii_lowercase
                 img_str = ''.join(random.choice(letters) for i in range(characters))
                 fullname[k] = str(img_str) + '.jpg'
                 filepath[k] = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + fullname[k]
                 urllib.request.urlretrieve(Image[k],filepath[k])
                 imagename.append(fullname[k])
                 print(fullname[k])
                 k=k+1
             j=j+1
         data.append((val,imagename))
         import pandas as pd
         df = pd.DataFrame(data,columns =['Link','Image1'])
         df.to_csv('E:\\Crawling-py-Data\\extractor\\New folder\\file4.csv',index=False,encoding='utf-8')
