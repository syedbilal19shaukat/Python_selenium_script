import urllib.request
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from random import randint
from boltons import iterutils
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
wk = openpyxl.load_workbook(r'C:\Users\Bilal Shaukat\Desktop\beauti-indution.xlsx')
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column
data = []
driver = webdriver.Chrome()
for i in range (1, rows+1):
         SKU_value = sh.cell(i,1).value
         val = iterutils.chunked(SKU_value, 128)
         print(val)
         driver.get('https://www.google.com/')
         sleep(randint(5,10))
         search_input = driver.find_element_by_xpath('//*[@id="tsf"]/div[2]/div/div[1]/div/div[1]/input').send_keys(val)
         search = driver.find_element_by_xpath('//*[@id="tsf"]/div[2]/div/div[1]/div/div[1]/input').send_keys( u'\ue007')
         time.sleep(2)
         try:
             Link1 = driver.find_element_by_xpath('//*[@id="rso"]/div[1]/div/div/div/div/div[1]/a[contains(@href, "https://www.amazon.com")]').get_attribute('href')
             Link_open = driver.find_element_by_xpath('//*[@id="rso"]/div[1]/div/div/div/div/div[1]/a[contains(@href, "https://www.amazon.com")]').click()
             print(Link1)
             time.sleep(4)
         except:
             try:
                 Link1 = driver.find_element_by_xpath('//*[@id="rso"]/div/div/div/div/div/div[1]/a[contains(@href,"https://www.amazon.com")] | //*[@id="rso"]/div[1]/div/div/div/div[1]/div/a[contains(@href, "https://www.amazon.com")]').get_attribute('href')
                 Link_open = driver.find_element_by_xpath('//*[@id="rso"]/div/div/div/div/div/div[1]/a[contains(@href,"https://www.amazon.com")] | //*[@id="rso"]/div[1]/div/div/div/div[1]/div/a[contains(@href, "https://www.amazon.com")]').click()
                 print(Link1)
                 time.sleep(4)
             except:
                 Link1 = 'N/A'
                 print(Link1)
         try:
             Paperback = driver.find_element_by_xpath('//td[@class="bucket"]/div[@class="content"]/ul/li[contains(b, "Paperback:")]').text
             print(Paperback)
         except:
             Paperback = 'N/A'
             print(Paperback)
         try:
             Publisher = driver.find_element_by_xpath('//td[@class="bucket"]/div[@class="content"]/ul/li[contains(b, "Publisher:")]').text
             print(Publisher)
         except:
             Publisher = 'N/A'
             print(Publisher)
         try: 
            Language = driver.find_element_by_xpath('//td[@class="bucket"]/div[@class="content"]/ul/li[contains(b, "Language:")]').text
            print(Language)
         except:
             Language = 'N/A'
             print(Language)
         try:
             ISBN_13 = driver.find_element_by_xpath('//td[@class="bucket"]/div[@class="content"]/ul/li[contains(b, "ISBN-13:")]').text
             print(ISBN_13)
         except:
             ISBN_13 = 'N/A'
             print(ISBN_13)
         try:
             Product_Dimensions = driver.find_element_by_xpath('//td[@class="bucket"]/div[@class="content"]/ul/li[contains(b, "Product Dimensions:")]').text
             print(Product_Dimensions)
         except:
             Product_Dimensions = 'N/A'
             print(Product_Dimensions)
         try:
             Shipping_Weight = driver.find_element_by_xpath('//td[@class="bucket"]/div[@class="content"]/ul/li[contains(b, "Shipping Weight:")]').text
             print(Shipping_Weight)
         except:
             Shipping_Weight = 'N/A'
             print(Shipping_Weight)
##         try:
##             categroy = driver.find_element_by_xpath('//ul[@class="items"]/li[2]').text
##             print(categroy)
##         except:
##             categroy = 'N/A'
##             print(categroy)
##         try:
##             brands = driver.find_element_by_xpath('//div[@class="additional-attributes-wrapper table-wrapper"]/table/tbody/tr[1]/td').text
##             print(brands)
##         except:
##             brands = 'N/A'
##             print(brands)
##         try:
##            Image2 = driver.find_element_by_xpath('//div[@class="prolabels-wrapper"]/img').get_attribute('src')
##            print(Image2)
##            len1 = 10
##            letters2 = string.ascii_lowercase
##            img2_name = ''.join(random.choice(letters2) for i in range(len1)) 
##            full2_name = str(img2_name) + '.jpg'
##            file2_path = 'E:\\crawlingold\\img\\' + full2_name
##            urllib.request.urlretrieve(Image2,file2_path)
##         except:
##            Image2 = 'No Image'
##            full2_name = 'No Image'
##            print(Image2)
         """try:             
             Manufacturer_Description = driver.find_element_by_xpath('//*[@id="aplus"]/div').text
             print(Manufacturer_Description)
         except:
             Manufacturer_Description = 'N/A'
             print(Manufacturer_Description)
         Image = dict()
         sampleImage = dict()
         fullname = dict()
         filepath = dict()
         j=1
         imagename=[]
         for images in driver.find_elements_by_xpath('//*[@id="altImages"]/ul/li[contains(@class,"imageThumbnail")]'):
             sampleImage[j] = driver.find_element_by_xpath('//*[@id="altImages"]/ul/li[contains(@class,"imageThumbnail")][%d]//img' % (j,)).get_attribute('src')
             Image[j] = sampleImage[j].replace('_SS40_','_SL1500_')
             print(Image[j])
             characters = 10
             letters = string.ascii_lowercase
             img_str = ''.join(random.choice(letters) for i in range(characters))
             fullname[j] = str(img_str) + '.jpg'
             filepath[j] = 'E:\\crawling\\AMAZON-DATA\\IMG\\' + fullname[j]
             urllib.request.urlretrieve(Image[j],filepath[j])
             imagename.append(fullname[j])
             print(fullname[j])
             j=j+1"""
         data.append((Link1,Paperback,Publisher,Language,ISBN_13,Product_Dimensions,Shipping_Weight))
         import pandas as pd
         df = pd.DataFrame(data,columns =['Link','Paperback','Publisher','Language','ISBN_13','Product_Dimensions','Shipping_Weight'])
         df.to_csv('E:\\Crawling-py-Data\\extractor\\New folder\\file2.csv',index=False,encoding='utf-8')
         
