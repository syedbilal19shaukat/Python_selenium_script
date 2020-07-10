import urllib.request
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
wk = openpyxl.load_workbook(r'C:\Users\Bilal Shaukat\Desktop\beauti-indution.xlsx')
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column
data = []
driver = webdriver.Chrome()
driver.maximize_window()
for i in range (1, rows+1):
         val = sh.cell(i,1).value
         driver.get(str(val))
         time.sleep(10)

         try:
            Image2 = driver.find_element_by_xpath('//*[@id="maincontent"]/div[3]/div[1]/div[4]/div[2]/div[2]/div[2]/div[1]/div[2]/div[1]/img').get_attribute('src')
            print(Image2)
            len2 = 10
            letters2 = string.ascii_lowercase
##            Image2 = Image2.replace('thumbnail','image').replace('120x150','1800x').replace('9df78eab33525d08d6e5fb8d27136e95','040ec09b1e35df139433887a97daa66f')
            img2_name = ''.join(random.choice(letters2) for i in range(len2))
            full2_name = str(img2_name) + '.jpg'
            file2_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full2_name
            urllib.request.urlretrieve(Image2,file2_path)
         except:
            Image2 = 'No Image'
            full2_name = 'No Image'
            print(Image2)
            pass
##         try:
##            Image3 = driver.find_element_by_xpath('//*[@id="product_addtocart_form"]/div[2]/div/div[2]/div/div[1]/ul/li[4]/a/img').get_attribute('src')
##            print(Image3)
##            len2 = 10
##            letters2 = string.ascii_lowercase
##            Image3 = Image3.replace('thumbnail','image').replace('120x150','1800x').replace('9df78eab33525d08d6e5fb8d27136e95','040ec09b1e35df139433887a97daa66f')
##            img3_name = ''.join(random.choice(letters2) for i in range(len2))
##            full3_name = str(img3_name) + '.jpg'
##            file3_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full3_name
##            urllib.request.urlretrieve(Image3,file3_path)
##         except:
##            Image3 = 'No Image'
##            full3_name = 'No Image'
##            print(Image3)
##            pass

         data.append((val,full2_name))
         import pandas as pd
         df = pd.DataFrame(data,columns =['PageURL','Image2'])
         df.to_csv('E:\\Crawling-py-Data\\extractor\\New folder\\tanishkimg.csv',index=False,encoding='utf-8')
         
           
        
            
          
           
            
            
            
            

            
                        
         
        

   



