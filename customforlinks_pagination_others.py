#----------- daraz crawl
import openpyxl
import sys, os, time
import re
import requests
from random import *
import random,string
from array import array
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
if __name__ == "__main__":
    if getattr(sys, 'frozen', False):
        chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")
        driver = webdriver.Chrome(chromedriver_path)
    else:
        driver = webdriver.Chrome()
path=input("Please Enter folder path: ")
os.chdir(path)
filename=input("Please Enter file name: ")
pagename=input('Enter store link: ')
pages1=input('Enter starting page: ')
page1=int(pages1)
pages2=input('Enter final page: ')
page2=int(pages2)
wb=openpyxl.Workbook()
sheet=wb.active
sheet.cell(row=1, column=1).value="Links"
print("Extract store link starts...")
x=0
urls=[]
browser = webdriver.Chrome()
for i in range(page1,page2+1):
    p=pagename.find('samsung-galaxy-a10s')
    if(p>1):
        newpage=pagename.replace('samsung-galaxy-a10s','samsung-galaxy-a10s?page='+str(i))
        browser.get(newpage)
        time.sleep(3)
    else:
        browser.get(pagename+"?page="+str(i)) 
        time.sleep(3)
    for j in range(1,501):
        k=str(j)
        d='//div[@class="row  categoryProduct xsResponse clearfix"]/div['    
        e=']/div/div/a'
        f=d+k+e
        try:
            dpage=browser.find_element_by_xpath(f).get_attribute('href')
        except:
            pass
        urls.insert(0,str(dpage))
urls=list(set(urls))
browser.quit()
for i in range(len(urls)):
    print("Link: "+str(i+1))
    sheet.cell(row=x+2, column=1).value=urls[i]
    x=x+1
    wb.save(filename+'.xlsx')
wb.save(filename+'.xlsx')
print("Download Compeltes...")
