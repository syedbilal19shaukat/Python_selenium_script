import urllib.request
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
wk = openpyxl.load_workbook(r'C:\Users\Bilal\Desktop\Zameen-listing.xlsx')
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column
data = []
driver = webdriver.Chrome()
driver.maximize_window()
driver.get('https://www.zameen.com/Property/gulshan_e_maymar_gulshan_e_maymar_sector_y_west_open_flat_is_available_for_sale-20733033-6761-1.html')
time.sleep(4)
for i in range (1, rows+1):
         val = sh.cell(i,1).value
         driver.get(val)
         time.sleep(10)
         try:
             title = driver.find_element_by_xpath('//*[@id="body-wrapper"]/main/div[2]/div/div[1]/h1').text
             print(title)
             time.sleep(2)
         except:
             title = 'No title'
             print(title)
         try:
             price = driver.find_element_by_xpath('//*[@id="body-wrapper"]/main/div[4]/div[1]/div[2]/div[1]/div[1]/div/div').text
             print(price)
             time.sleep(2)
         except:
             price = 'No price'
             print(price)
         try:
             mypd = browser.find_element_by_tag_name('html')
             mypd.send_keys(Keys.DOWN)
             time.sleep(5)
         except:
            pass
         try:
             mypd2 = browser.find_element_by_tag_name('body')
             mypd2.send_keys(Keys.DOWN)
             time.sleep(5)
         except:
            pass
         try:
             mypd3 = browser.find_element_by_tag_name('header')
             mypd3.send_keys(Keys.DOWN)
             time.sleep(5)
         except:
            pass
         try:
             mypd4 = browser.find_element_by_tag_name('main')
             mypd4.send_keys(Keys.DOWN)
             time.sleep(5)
         except:
            pass
         try:
             mypd3 = browser.find_element_by_tag_name('h3')
             mypd3.send_keys(Keys.DOWN)
             time.sleep(5)
         except:
            pass
         try:
             mypd3 = browser.find_element_by_tag_name('span')
             mypd3.send_keys(Keys.DOWN)
             time.sleep(5)
         except:
            pass
         try:
             mypd4 = browser.find_element_by_tag_name('footer')
             mypd4.send_keys(Keys.DOWN)
             time.sleep(5)
         except:
            pass
         try:
             mypd3 = browser.find_element_by_tag_name('div')
             mypd3.send_keys(Keys.DOWN)
             time.sleep(5)
         except:
            pass
         read_more = driver.find_element_by_xpath('//*[@id="body-wrapper"]/main/div[4]/div[1]/div[4]/div/div[3]/div[2]').click()
         time.sleep(2)
         try:
             Description = driver.find_element_by_xpath('//div[@class="_1aca585a c352c124"]/div[3]/div/div/div/div/span').text
             print(Description)
             time.sleep(2)
         except:
             Description = 'No Description'
             print(Description)
         image_picture = driver.find_element_by_xpath('//*[@id="body-wrapper"]/main/div[4]/div[1]/div[1]/div[1]/div/div[1]/div/div[1]').click()
         time.sleep(2)
         try:
            Image1 = driver.find_element_by_xpath('//div[@class="image-gallery-thumbnails-wrapper bottom "]/div[@class="image-gallery-thumbnails"]/div[@class="image-gallery-thumbnails-container"]/a[@class="image-gallery-thumbnail"][1]/picture/img').get_attribute('src')
            print(Image1)
##            len1 = 10
##            letters1 = string.ascii_lowercase
##            full1_name=Image1.replace('-120x90','-800x600')
##            img1_name = ''.join(random.choice(letters1) for i in range(len1))
##            full1_name = str(img1_name) + '.jpg'
##            file1_path = 'D:\\Work\\New-Folder\\Zameen.com\\' + full1_name
##            urllib.request.urlretrieve(Image1,file1_path)
         except:
##            Image1 = 'No Image'
##            full1_name = 'No Image'
##            print(Image1)
            pass
         try:
            Image2 = driver.find_element_by_xpath('//div[@class="image-gallery-thumbnails-wrapper bottom "]/div[@class="image-gallery-thumbnails"]/div[@class="image-gallery-thumbnails-container"]/a[@class="image-gallery-thumbnail"][2]/picture/img').get_attribute('src')
            print(Image2)
##            len1 = 10
##            letters1 = string.ascii_lowercase
##            Image2=Image2.replace('-120x90','-800x600')
##            img2_name = ''.join(random.choice(letters1) for i in range(len1))
##            full2_name = str(img2_name) + '.jpg'
##            file2_path = 'D:\\Work\\New-Folder\\Zameen.com\\' + full2_name
##            urllib.request.urlretrieve(Image2,file2_path)
         except:
##            Image2 = 'No Image'
##            full2_name = 'No Image'
##            print(Image2)
            pass
         try:
            Image3 = driver.find_element_by_xpath('//div[@class="image-gallery-thumbnails-wrapper bottom "]/div[@class="image-gallery-thumbnails"]/div[@class="image-gallery-thumbnails-container"]/a[@class="image-gallery-thumbnail"][3]/picture/img').get_attribute('src')
            print(Image3)
##            len1 = 10
##            letters1 = string.ascii_lowercase
##            Image3=Image3.replace('-120x90','-800x600')
##            img3_name = ''.join(random.choice(letters1) for i in range(len1))
##            full3_name = str(img3_name) + '.jpg'
##            file3_path = 'D:\\Work\\New-Folder\\Zameen.com\\' + full3_name
##            urllib.request.urlretrieve(Image3,file3_path)
         except:
##            Image3 = 'No Image'
##            full3_name = 'No Image'
##            print(Image3)
            pass

         try:
            Image4 = driver.find_element_by_xpath('//div[@class="image-gallery-thumbnails-wrapper bottom "]/div[@class="image-gallery-thumbnails"]/div[@class="image-gallery-thumbnails-container"]/a[@class="image-gallery-thumbnail"][4]/picture/img').get_attribute('src')
            print(Image4)
##            len1 = 10
##            letters1 = string.ascii_lowercase
##            Image4=Image4.replace('-120x90','-800x600')
##            img4_name = ''.join(random.choice(letters1) for i in range(len1))
##            full4_name = str(img4_name) + '.jpg'
##            file4_path = 'D:\\Work\\New-Folder\\Zameen.com\\' + full4_name
##            urllib.request.urlretrieve(Image4,file4_path)
         except:
##            Image4 = 'No Image'
##            full4_name = 'No Image'
##            print(Image4)
            pass
         try:
            Image5 = driver.find_element_by_xpath('//div[@class="image-gallery-thumbnails-wrapper bottom "]/div[@class="image-gallery-thumbnails"]/div[@class="image-gallery-thumbnails-container"]/a[@class="image-gallery-thumbnail"][5]/picture/img').get_attribute('src')
            print(Image5)
##            len1 = 10
##            letters1 = string.ascii_lowercase
##            Image5 = Image5.replace('-120x90','-800x600')
##            img5_name = ''.join(random.choice(letters1) for i in range(len1))
##            full5_name = str(img5_name) + '.jpg'
##            file5_path = 'D:\\Work\\New-Folder\\Zameen.com\\' + full5_name
##            urllib.request.urlretrieve(Image5,file5_path)
         except:
##            Image5 = 'No Image'
##            full5_name = 'No Image'
##            print(Image5)
            pass
         try:
            Image6 = driver.find_element_by_xpath('//div[@class="image-gallery-thumbnails-wrapper bottom "]/div[@class="image-gallery-thumbnails"]/div[@class="image-gallery-thumbnails-container"]/a[@class="image-gallery-thumbnail"][6]/picture/img').get_attribute('src')
            print(Image6)
##            len1 = 10
##            letters1 = string.ascii_lowercase
##            Image6 = Image6.replace('-120x90','-800x600')
##            img6_name = ''.join(random.choice(letters1) for i in range(len1))
##            full6_name = str(img6_name) + '.jpg'
##            file6_path = 'D:\\Work\\New-Folder\\Zameen.com\\' + full6_name
##            urllib.request.urlretrieve(Image6,file6_path)
         except:
##            Image6 = 'No Image'
##            full6_name = 'No Image'
##            print(Image6)
            pass
         try:
            Image7 = driver.find_element_by_xpath('//div[@class="image-gallery-thumbnails-wrapper bottom "]/div[@class="image-gallery-thumbnails"]/div[@class="image-gallery-thumbnails-container"]/a[@class="image-gallery-thumbnail"][7]/picture/img').get_attribute('src')
            print(Image7)
##            len1 = 10
##            letters1 = string.ascii_lowercase
##            Image7 = Image7.replace('-120x90','-800x600')
##            img7_name = ''.join(random.choice(letters1) for i in range(len1))
##            full7_name = str(img7_name) + '.jpg'
##            file7_path = 'D:\\Work\\New-Folder\\Zameen.com\\' + full7_name
##            urllib.request.urlretrieve(Image7,file7_path)
         except:
##            Image7 = 'No Image'
##            full7_name = 'No Image'
##            print(Image7)
            pass
         try:
            Image8 = driver.find_element_by_xpath('//div[@class="image-gallery-thumbnails-wrapper bottom "]/div[@class="image-gallery-thumbnails"]/div[@class="image-gallery-thumbnails-container"]/a[@class="image-gallery-thumbnail"][8]/picture/img').get_attribute('src')
            print(Image8)
##            len1 = 10
##            letters1 = string.ascii_lowercase
##            Image8 = Image8.replace('-120x90','-800x600')
##            img8_name = ''.join(random.choice(letters1) for i in range(len1))
##            full8_name = str(img8_name) + '.jpg'
##            file8_path = 'D:\\Work\\New-Folder\\Zameen.com\\' + full8_name
##            urllib.request.urlretrieve(Image8,file8_path)
         except:
##            Image8 = 'No Image'
##            full8_name = 'No Image'
##            print(Image8)
            pass
        
         try:
            Image9 = driver.find_element_by_xpath('//div[@class="image-gallery-thumbnails-wrapper bottom "]/div[@class="image-gallery-thumbnails"]/div[@class="image-gallery-thumbnails-container"]/a[@class="image-gallery-thumbnail"][9]/picture/img').get_attribute('src')
            print(Image9)
##            len1 = 10
##            letters1 = string.ascii_lowercase
##            Image9 = Image9.replace('-120x90','-800x600')
##            img9_name = ''.join(random.choice(letters1) for i in range(len1))
##            full9_name = str(img9_name) + '.jpg'
##            file9_path = 'D:\\Work\\New-Folder\\Zameen.com\\' + full9_name
##            urllib.request.urlretrieve(Image9,file9_path)
         except:
##            Image9 = 'No Image'
##            full9_name = 'No Image'
##            print(Image9)
            pass
         time.sleep(2)         
         image_modal_close = driver.find_element_by_xpath('//button[@class="eae3ed68 c1c2e5c8"]').click()
         time.sleep(2)         
         image_button_open = driver.find_element_by_xpath('//*[@id="body-wrapper"]/main/div[4]/div[2]/div[1]/div/form/div[2]/button[1]').click()
         time.sleep(2)
         try:
             Mob_number = driver.find_element_by_xpath('//div[@class="_2ff591d9"]/table/tbody/tr[1]/td[2]/a').text
             print(Mob_number)
         except:
             Mob_number = 'No Mob_number'
             print(Mob_number)
             pass
         try:
             phone_number = driver.find_element_by_xpath('//div[@class="_2ff591d9"]/table/tbody/tr[2]/td[2]/a').text
             print(phone_number)
         except:
             phone_number = 'No phone_number'
             print(phone_number)
             pass
         try:
             Agency_Name = driver.find_element_by_xpath('//div[@class="_2ff591d9"]/span[2]').text
             print(Agency_Name)
         except:
             Agency_Name = 'No Agency_Name'
             print(Agency_Name)
             pass
         try:
             Agent_name = driver.find_element_by_xpath('//div[@class="_2ff591d9"]/table/tbody/tr[3]/td[2]').text
             print(Agent_name)
         except:
             Agent_name = 'No Agent_name'
             print(Agent_name)
             pass
         time.sleep(2)         
         image_modal_close = driver.find_element_by_xpath('//div[@class="_2ff591d9"]/button').click()
         time.sleep(2)
        
         data.append((val,title,price,Description,Image1,Image2,Image3,Image4,Image5,Image6,Image7,Image8,Image9,Mob_number,phone_number,Agency_Name,Agent_name))
         import pandas as pd
         df = pd.DataFrame(data,columns =['Link','Title','price','Description','Image1','Image2','Image3','Image4','Image5','Image6','Image7','Image8','Image9','Mobile-Number','Phone-Number','Agency-Name','Agent-Name'])
         df.to_csv('D:\\Work\\New-Folder\\Zameen.com\\geonline.csv',index=False,encoding='utf-8')
         
           
##         try:   
##             stock = driver.find_element_by_xpath('//div[@class="product-stats"]/ul/li/b[contains(text(),"Stock:")]').text
##             print(stock)
##         except:
##             stock = 'No stock'
##             print(stock)
##         try:   
##             weight = driver.find_element_by_xpath('//div[@class="product-stats"]/ul/li/b[contains(text(),"Weight:")]').text
##             print(weight)
##         except:
##             weight = 'No weight'
##             print(weight)
##         try:
##            Image1 = driver.find_element_by_xpath('/html/body/app-root/div[1]/app-base/mat-sidenav-container/mat-sidenav-content/div/app-product/div/div[1]/div[2]/app-product-images/div/div/div/div[2]/div[1]/div[1]/div/img').get_attribute('src')
##            print(Image1)
##            len1 = 10
##            letters1 = string.ascii_lowercase
##            img1_name = ''.join(random.choice(letters1) for i in range(len1))
##            full1_name = str(img1_name) + '.jpg'
##            file1_path = 'D:\\Work\\New-Folder\\Zameen.com\\' + full1_name
##            urllib.request.urlretrieve(Image1,file1_path)
##         except:
##            Image1 = 'No Image'
##            full1_name = 'No Image'
##            print(Image1)
##            pass
##         try:
##            Image2 = driver.find_element_by_xpath('/html/body/app-root/div[1]/app-base/mat-sidenav-container/mat-sidenav-content/div/app-product/div/div[1]/div[2]/app-product-images/div/div/div/div[2]/div[1]/div[2]/div/img').get_attribute('src')
##            print(Image2)
##            len1 = 10
##            letters1 = string.ascii_lowercase
##            img2_name = ''.join(random.choice(letters1) for i in range(len1))
##            full2_name = str(img2_name) + '.jpg'
##            file2_path = 'D:\\Work\\New-Folder\\Zameen.com\\' + full2_name
##            urllib.request.urlretrieve(Image2,file2_path)
##         except:
##            Image2 = 'No Image'
##            full2_name = 'No Image'
##            print(Image2)
##            pass
##         try:
##            Image3 = driver.find_element_by_xpath('//*[@id="form_singleProduct"]/div/div/div/div[1]/div/div[2]/div[2]/div/div[2]/div[1]/div/div[9]/div/img').get_attribute('src')
##            print(Image3)
##            len1 = 10
##            letters1 = string.ascii_lowercase
##            img3_name = ''.join(random.choice(letters1) for i in range(len1))
##            full3_name = str(img3_name) + '.jpg'
##            file3_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full3_name
##            urllib.request.urlretrieve(Image3,file3_path)
##         except:
##            Image3 = 'No Image'
##            full3_name = 'No Image'
##            print(Image3)
##            pass
##         try:
##            Image4 = driver.find_element_by_xpath('//*[@id="form_singleProduct"]/div/div/div/div[1]/div/div[2]/div[2]/div/div[2]/div[1]/div/div[6]/div/img').get_attribute('src')
##            print(Image4)
##            len1 = 10
##            letters1 = string.ascii_lowercase
##            img4_name = ''.join(random.choice(letters1) for i in range(len1))
##            full4_name = str(img4_name) + '.jpg'
##            file4_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full4_name
##            urllib.request.urlretrieve(Image4,file4_path)
##         except:
##            Image4 = 'No Image'
##            full4_name = 'No Image'
##            print(Image4)
##            pass
        
            
          
           
            
            
            
            

            
                        
         
        

   



