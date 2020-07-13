import sys, os, time
import urllib.request
import getpass
import py2exe
import requests
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
##wk = openpyxl.load_workbook(r'C:\Users\Bilal\Desktop\test.xlsx')
##sh = wk['Sheet1']
##rows = sh.max_row
##col = sh.max_colum
CSV_path1 = input("Please enter location to save file: ")
CSV_path =  CSV_path1.replace('/','//')
data = []
if getattr(sys, 'frozen', False):
    chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")
    driver = webdriver.Chrome(chromedriver_path)
else:
    driver = webdriver.Chrome()
driver.maximize_window()
for i in range (1, 2):
##         val = sh.cell(i,1).value
         driver.get('https://www.theroxycinemas.com/movies')
         time.sleep(5)
         try:      
             read_more = driver.find_element_by_xpath('//*[@id="crckies"]').click()
             time.sleep(3)
         except:
             pass
         j=1
         imagename=[]
         skuu=[] 
         for link in driver.find_elements_by_xpath('//*[@id="ulnowshowing"]/li'): 
             time.sleep(6)
             link = driver.find_element_by_xpath('//*[@id="ulnowshowing"]/li[%d]' % (j,)).click()
             link2 = driver.current_url
             print(link2)
             Image = dict()
             sampleImage = dict()
             fullname = dict()
             filepath = dict()
             time.sleep(4)
             try:
                 title = driver.find_element_by_xpath('//*[@id="MainContainer"]/section[1]/section/section/section[1]/section/section/section[1]/section/h1').text
                 print(title)
                 time.sleep(2)
             except:
                    title = 'No title'
                    print(title)
             try:
                 language = driver.find_element_by_xpath('//*[@id="MainContainer"]/section[1]/section/section/section[1]/section/section/section[1]/section/section[1]/section[2]/aside/ul/li[1]/p').text
                 print(language)
                 time.sleep(2)
             except:
                    language = 'No language'
                    print(language)
             try:       
                 read_more = driver.find_element_by_xpath('//*[@id="MainContainer"]/section[1]/section/section/section[1]/section/section/section[1]/section/section[3]/p/a').click()
                 time.sleep(3)
             except:
                 pass
##             try:       
##                 if driver.find_element_by_xpath('//*[@id="MainContainer"]/section[1]/section/section/section[1]/section/section/section[1]/section/section[3]/p'):
##                     description = driver.find_element_by_xpath('//*[@id="MainContainer"]/section[1]/section/section/section[1]/section/section/section[1]/section/section[3]/p').text
##                     print(description)
##                     time.sleep(2)
##                 else:
##                     description_without_readmore = driver.find_element_by_xpath('//*[@id="MainContainer"]/section[1]/section/section/section[1]/section/section/section[1]/section/section[3]/p').text
##                     print(description_without_readmore)
##                     time.sleep(2)
##             except:
##                 pass                     
##             try:
##                 description = driver.find_element_by_xpath('//*[@id="MainContainer"]/section[1]/section/section/section[1]/section/section/section[1]/section/section[3]/p').text
##                 print(description)
##                 time.sleep(2)
##             except:
##                    description = 'No description'
##                    print(description)
##                    pass
             try:
                 description_without_readmore = driver.find_element_by_xpath('//*[@id="MainContainer"]/section[1]/section/section/section[1]/section/section/section[1]/section/section[3]/p').text
                 print(description_without_readmore)
                 time.sleep(2)
             except:
                    description_without_readmore = 'No description_without_readmore'
                    print(description_without_readmore)
             driver.get('https://www.theroxycinemas.com/movies')   
             j=j+1
             data.append((link2,title,language,description_without_readmore))
             import pandas as pd
             df = pd.DataFrame(data,columns =['url','title','language','description_without_readmore'])
             df.to_csv(CSV_path+'//file.csv',index=False,encoding='utf-8')
