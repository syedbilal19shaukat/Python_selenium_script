import urllib.request
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from random import randint
from boltons import iterutils
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
wk = openpyxl.load_workbook(r'C:\Users\Bilal Shaukat\Desktop\beauti-indution.xlsx')
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column
data = []
driver = webdriver.Chrome()
for i in range (1, rows+1):
         SKU_value = sh.cell(i,1).value
         val = iterutils.chunked(SKU_value, 128)
         driver.get('https://www.goto.com.pk')
         time.sleep(6)
         driver.get('https://www.gsmarena.com')
         time.sleep(6) 
         print(val)
         driver.get('https://www.google.com/')
         sleep(randint(15,50))
         search_input = driver.find_element_by_xpath('//*[@id="tsf"]/div[2]/div/div[1]/div/div[1]/input').send_keys(val)
         search = driver.find_element_by_xpath('//*[@id="tsf"]/div[2]/div/div[1]/div/div[1]/input').send_keys( u'\ue007')
         time.sleep(6)
         try:
             Link1 = driver.find_element_by_xpath('//*[@id="rso"]/div[1]/div/div/div/div/div[1]/a[contains(@href, "https://homeshopping.pk")]').get_attribute('href')
             Link_open = driver.find_element_by_xpath('//*[@id="rso"]/div[1]/div/div/div/div/div[1]/a[contains(@href, "https://homeshopping.pk")]').click()
             print(Link1)
             time.sleep(10)
         except:
             try:
                 Link1 = driver.find_element_by_xpath('//*[@id="rso"]/div/div/div/div/div/div[1]/a[contains(@href,"https://homeshopping.pk")] | //*[@id="rso"]/div[1]/div/div/div/div[1]/div/a[contains(@href, "https://homeshopping.pk")]').get_attribute('href')
                 Link_open = driver.find_element_by_xpath('//*[@id="rso"]/div/div/div/div/div/div[1]/a[contains(@href,"https://homeshopping.pk")] | //*[@id="rso"]/div[1]/div/div/div/div[1]/div/a[contains(@href, "https://homeshopping.pk")]').click()
                 print(Link1)
                 time.sleep(10)
             except:
                 Link1 = 'N/A'
                 print(Link1)
         data.append((Link1,val))
         import pandas as pd
         df = pd.DataFrame(data,columns =['Link','Name'])
         df.to_csv('E:\\Crawling-py-Data\\extractor\\New folder\\newfile5.csv',index=False,encoding='utf-8')
         driver.get('https://www.youtube.com')
 
