import urllib.request
import urllib.parse
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import Select
from bs4 import BeautifulSoup
import re
import requests
wk = openpyxl.load_workbook(r'C:\Users\Bilal Shaukat\Desktop\beauti-indution.xlsx')
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column     #//div[@class="swatch clearfix"]/div[@data-value="CBL101 Mind Reader"]/label/@style
data = []
driver = webdriver.Chrome()
driver.maximize_window()
for i in range (1, rows+1):
         val = sh.cell(i,1).value
         driver.get(val)
         time.sleep(10)
         try:
             Image2 = driver.find_element_by_xpath('//div[@class="swatch clearfix"]/div[@data-value="CBL101 Mind Reader"]/label[contains(@style, "background-image: url(https://cdn.shopify.com/s/files/1/0072/4847/8306/t/4/assets")]').get_attribute('style')
             print(Image2)
             len1 = 10
             letters2 = string.ascii_lowercase
             Image2=Image2.replace("background-image: url(\"","").replace("\");","")
             #Image2=Image2.replace("\");","")
             print(Image2)
             img2_name = ''.join(random.choice(letters2) for i in range(len1)) 
             full2_name = str(img2_name) + '.jpg'
             file2_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full2_name
             urllib.request.urlretrieve(Image2,file2_path)
         except:
             Image2 = 'No Image'
             print(Image2)
             pass
             
         
         data.append((full2_name,val))
import pandas as pd
df = pd.DataFrame(data,columns =['Image2','PageURL'])
df.to_csv('E:\\Crawling-py-Data\\extractor\\New folder\\file.csv',index=False,encoding='utf-8')
