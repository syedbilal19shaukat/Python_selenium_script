import urllib.request
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from random import randint
from boltons import iterutils
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
wk = openpyxl.load_workbook(r'C:\Users\Bilal Shaukat\Desktop\beauti-indution.xlsx')
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column
data = []
driver = webdriver.Chrome()
for i in range (1, rows+1):
         SKU_value = sh.cell(i,1).value
         val = iterutils.chunked(SKU_value, 128)
         print(val)
         driver.get('https://search.yahoo.com/')
         sleep(randint(2,6))
         search_input = driver.find_element_by_xpath('//*[@id="yschsp"]').send_keys(val)
         search = driver.find_element_by_xpath('//*[@id="yschsp"]').send_keys( u'\ue007')
         time.sleep(2)
         try:
             Link1 = driver.find_element_by_xpath('//ol[@class="mb-15 reg searchCenterMiddle"]/li/div/div/h3/a[contains(@href, "https://www.samsung.com")]').get_attribute('href')
             Link_open = driver.find_element_by_xpath('//ol[@class="mb-15 reg searchCenterMiddle"]/li/div/div/h3/a[contains(@href, "https://www.samsung.com")]').click()
             print(Link1)
             time.sleep(4)
         except:
             try:
                 Link1 = driver.find_element_by_xpath('//ol[@class="mb-15 reg searchCenterMiddle"]/li/div/div/h3/a[contains(@href, "https://www.samsung.com")] | //ol[@class="mb-15 reg searchCenterMiddle"]/li/div/div/h3/a[contains(@href, "https://www.samsung.com")]').get_attribute('href')
                 Link_open = driver.find_element_by_xpath('//ol[@class="mb-15 reg searchCenterMiddle"]/li/div/div/h3/a[contains(@href, "https://www.samsung.com")] | //ol[@class="mb-15 reg searchCenterMiddle"]/li/div/div/h3/a[contains(@href, "https://www.samsung.com")]').click()
                 print(Link1)
                 time.sleep(4)
             except:
                 Link1 = 'N/A'
                 print(Link1)
         data.append((Link1,val))
         import pandas as pd
         df = pd.DataFrame(data,columns =['Link','Name'])
         df.to_csv('E:\\Crawling-py-Data\\extractor\\New folder\\file.csv',index=False,encoding='utf-8')
 
