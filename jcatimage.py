import urllib.request
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
wk = openpyxl.load_workbook(r'C:\Users\Bilal Shaukat\Desktop\New.xlsx')
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column
data = []
driver = webdriver.Chrome()
driver.maximize_window()
for i in range (1, rows+1):
         val = sh.cell(i,1).value
         driver.get(val)
         time.sleep(10)
         try:
            Image2 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[1]/a/div/span/img[1]').get_attribute('src')
            print(Image2)
            len2 = 10
            letters2 = string.ascii_lowercase
            img2_name = ''.join(random.choice(letters2) for i in range(len2))
            full2_name = str(img2_name) + '.jpg'
            file2_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full2_name
            urllib.request.urlretrieve(Image2,file2_path)
         except:
            Image2 = 'No Image'
            print(Image2)
         try:
            Image3 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[2]/a/div/span/img[1]').get_attribute('src')
            print(Image3)
            len3 = 10
            letters3 = string.ascii_lowercase
            img3_name = ''.join(random.choice(letters3) for i in range(len3))
            full3_name = str(img3_name) + '.jpg'
            file3_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full3_name
            urllib.request.urlretrieve(Image3,file3_path)
         except:
            Image3 = 'No Image'
            print(Image3)
         try:
            Image4 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[3]/a/div/span/img[1]').get_attribute('src')
            print(Image4)
            len4 = 10
            letters4 = string.ascii_lowercase
            img4_name = ''.join(random.choice(letters4) for i in range(len4))
            full4_name = str(img4_name) + '.jpg'
            file4_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full4_name
            urllib.request.urlretrieve(Image4,file4_path)
         except:
            Image4 = 'No Image'
            print(Image4)
         try:
            Image5 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[4]/a/div/span/img[1]').get_attribute('src')
            print(Image5)
            len5 = 10
            letters5 = string.ascii_lowercase
            img5_name = ''.join(random.choice(letters5) for i in range(len5))
            full5_name = str(img5_name) + '.jpg'
            file5_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full5_name
            urllib.request.urlretrieve(Image5,file5_path)
         except:
            Image5 = 'No Image'
            print(Image5)
         try:
            Image6 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[5]/a/div/span/img[1]').get_attribute('src')
            print(Image6)
            len6 = 10
            letters6 = string.ascii_lowercase
            img6_name = ''.join(random.choice(letters6) for i in range(len6))
            full6_name = str(img6_name) + '.jpg'
            file6_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full6_name
            urllib.request.urlretrieve(Image6,file6_path)
         except:
            Image6 = 'No Image'
            print(Image6)
         try:
            Image7 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[6]/a/div/span/img[1]').get_attribute('src')
            print(Image7)
            len7 = 10
            letters7 = string.ascii_lowercase
            img7_name = ''.join(random.choice(letters7) for i in range(len7))
            full7_name = str(img7_name) + '.jpg'
            file7_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full7_name
            urllib.request.urlretrieve(Image7,file7_path)
         except:
            Image7 = 'No Image'
            print(Image7)
         try:
            Image8 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[7]/a/div/span/img[1]').get_attribute('src')
            print(Image8)
            len8 = 10
            letters8 = string.ascii_lowercase
            img8_name = ''.join(random.choice(letters8) for i in range(len8))
            full8_name = str(img8_name) + '.jpg'
            file8_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full8_name
            urllib.request.urlretrieve(Image8,file8_path)
         except:
            Image8 = 'No Image'
            print(Image8)
         try:
            Image9 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[8]/a/div/span/img[1]').get_attribute('src')
            print(Image9)
            len9 = 10
            letters9 = string.ascii_lowercase
            img9_name = ''.join(random.choice(letters9) for i in range(len9))
            full9_name = str(img9_name) + '.jpg'
            file9_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full9_name
            urllib.request.urlretrieve(Image9,file9_path)
         except:
            Image9 = 'No Image'
            print(Image9)
         try:
            Image10 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[9]/a/div/span/img[1]').get_attribute('src')
            print(Image10)
            len10 = 10
            letters10 = string.ascii_lowercase
            img10_name = ''.join(random.choice(letters10) for i in range(len10))
            full10_name = str(img10_name) + '.jpg'
            file10_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full10_name
            urllib.request.urlretrieve(Image10,file10_path)
         except:
            Image10 = 'No Image'
            print(Image10)
         try:
            Image12 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[12]/a/div/span/img[1]').get_attribute('src')
            print(Image12)
            len12 = 10
            letters12 = string.ascii_lowercase
            img12_name = ''.join(random.choice(letters12) for i in range(len12))
            full12_name = str(img12_name) + '.jpg'
            file12_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full12_name
            urllib.request.urlretrieve(Image12,file12_path)
         except:
            Image12 = 'No Image'
            print(Image12)
         try:
            Image13 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[13]/a/div/span/img[1]').get_attribute('src')
            print(Image13)
            len13 = 10
            letters13 = string.ascii_lowercase
            img13_name = ''.join(random.choice(letters13) for i in range(len13))
            full13_name = str(img13_name) + '.jpg'
            file13_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full13_name
            urllib.request.urlretrieve(Image13,file13_path)
         except:
            Image13 = 'No Image'
            print(Image13)
         try:
            Image14 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[14]/a/div/span/img[1]').get_attribute('src')
            print(Image14)
            len14 = 10
            letters14 = string.ascii_lowercase
            img14_name = ''.join(random.choice(letters14) for i in range(len14))
            full14_name = str(img14_name) + '.jpg'
            file14_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full14_name
            urllib.request.urlretrieve(Image14,file14_path)
         except:
            Image14 = 'No Image'
            print(Image14)
         try:
            Image15 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[15]/a/div/span/img[1]').get_attribute('src')
            print(Image15)
            len15 = 10
            letters15 = string.ascii_lowercase
            img15_name = ''.join(random.choice(letters15) for i in range(len15))
            full15_name = str(img15_name) + '.jpg'
            file15_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full15_name
            urllib.request.urlretrieve(Image15,file15_path)
         except:
            Image15 = 'No Image'
            print(Image15)
         try:
            Image16 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[16]/a/div/span/img[1]').get_attribute('src')
            print(Image16)
            len16 = 10
            letters16 = string.ascii_lowercase
            img16_name = ''.join(random.choice(letters16) for i in range(len16))
            full16_name = str(img16_name) + '.jpg'
            file16_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full16_name
            urllib.request.urlretrieve(Image16,file16_path)
         except:
            Image16 = 'No Image'
            print(Image16)
         try:
            Image17 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[17]/a/div/span/img[1]').get_attribute('src')
            print(Image17)
            len17 = 10
            letters17 = string.ascii_lowercase
            img17_name = ''.join(random.choice(letters17) for i in range(len17))
            full17_name = str(img17_name) + '.jpg'
            file17_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full17_name
            urllib.request.urlretrieve(Image17,file17_path)
         except:
            Image17 = 'No Image'
            print(Image17)
         try:
            Image18 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[18]/a/div/span/img[1]').get_attribute('src')
            print(Image18)
            len18 = 10
            letters18 = string.ascii_lowercase
            img18_name = ''.join(random.choice(letters18) for i in range(len18))
            full18_name = str(img18_name) + '.jpg'
            file18_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full18_name
            urllib.request.urlretrieve(Image18,file18_path)
         except:
            Image18 = 'No Image'
            print(Image18)
         try:
            Image19 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[19]/a/div/span/img[1]').get_attribute('src')
            print(Image19)
            len19 = 10
            letters19 = string.ascii_lowercase
            img19_name = ''.join(random.choice(letters19) for i in range(len19))
            full19_name = str(img19_name) + '.jpg'
            file19_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full19_name
            urllib.request.urlretrieve(Image19,file19_path)
         except:
            Image19 = 'No Image'
            print(Image19)
         try:
            Image20 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[20]/a/div/span/img[1]').get_attribute('src')
            print(Image20)
            len20 = 10
            letters20 = string.ascii_lowercase
            img20_name = ''.join(random.choice(letters20) for i in range(len20))
            full20_name = str(img5_name) + '.jpg'
            file20_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full20_name
            urllib.request.urlretrieve(Image20,file20_path)
         except:
            Image20 = 'No Image'
            print(Image20)
         try:
            Image21 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[21]/a/div/span/img[1]').get_attribute('src')
            print(Image21)
            len21 = 10
            letters21 = string.ascii_lowercase
            img21_name = ''.join(random.choice(letters21) for i in range(len21))
            full21_name = str(img21_name) + '.jpg'
            file21_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full21_name
            urllib.request.urlretrieve(Image21,file21_path)
         except:
            Image21 = 'No Image'
            print(Image21)
         try:
            Image22 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[22]/a/div/span/img[1]').get_attribute('src')
            print(Image22)
            len22 = 10
            letters22 = string.ascii_lowercase
            img22_name = ''.join(random.choice(letters22) for i in range(len22))
            full22_name = str(img22_name) + '.jpg'
            file22_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full22_name
            urllib.request.urlretrieve(Image22,file22_path)
         except:
            Image22 = 'No Image'
            print(Image22)
         try:
            Image23 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[23]/a/div/span/img[1]').get_attribute('src')
            print(Image23)
            len23 = 10
            letters23 = string.ascii_lowercase
            img23_name = ''.join(random.choice(letters23) for i in range(len23))
            full23_name = str(img23_name) + '.jpg'
            file23_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full23_name
            urllib.request.urlretrieve(Image23,file23_path)
         except:
            Image23 = 'No Image'
            print(Image23)
         try:
            Image24 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[24]/a/div/span/img[1]').get_attribute('src')
            print(Image24)
            len24 = 10
            letters24 = string.ascii_lowercase
            img24_name = ''.join(random.choice(letters24) for i in range(len24))
            full24_name = str(img24_name) + '.jpg'
            file24_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full24_name
            urllib.request.urlretrieve(Image24,file24_path)
         except:
            Image24 = 'No Image'
            print(Image24)
         try:
            Image25 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[25]/a/div/span/img[1]').get_attribute('src')
            print(Image25)
            len25 = 10
            letters25 = string.ascii_lowercase
            img25_name = ''.join(random.choice(letters25) for i in range(len25))
            full25_name = str(img25_name) + '.jpg'
            file25_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full25_name
            urllib.request.urlretrieve(Image25,file25_path)
         except:
            Image25 = 'No Image'
            print(Image25)
         try:
            Image26 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[26]/a/div/span/img[1]').get_attribute('src')
            print(Image26)
            len26 = 10
            letters26 = string.ascii_lowercase
            img26_name = ''.join(random.choice(letters26) for i in range(len26))
            full26_name = str(img26_name) + '.jpg'
            file26_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full26_name
            urllib.request.urlretrieve(Image26,file26_path)
         except:
            Image26 = 'No Image'
            print(Image26)
         try:
            Image27 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[27]/a/div/span/img[1]').get_attribute('src')
            print(Image27)
            len27 = 10
            letters27 = string.ascii_lowercase
            img27_name = ''.join(random.choice(letters27) for i in range(len27))
            full27_name = str(img27_name) + '.jpg'
            file27_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full27_name
            urllib.request.urlretrieve(Image27,file27_path)
         except:
            Image27 = 'No Image'
            print(Image27)
         try:
            Image28 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[28]/a/div/span/img[1]').get_attribute('src')
            print(Image28)
            len28 = 10
            letters28 = string.ascii_lowercase
            img28_name = ''.join(random.choice(letters28) for i in range(len28))
            full28_name = str(img28_name) + '.jpg'
            file28_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full28_name
            urllib.request.urlretrieve(Image28,file28_path)
         except:
            Image28 = 'No Image'
            print(Image28)
         try:
            Image29 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[29]/a/div/span/img[1]').get_attribute('src')
            print(Image29)
            len29 = 10
            letters29 = string.ascii_lowercase
            img29_name = ''.join(random.choice(letters29) for i in range(len29))
            full29_name = str(img29_name) + '.jpg'
            file29_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full29_name
            urllib.request.urlretrieve(Image29,file29_path)
         except:
            Image29 = 'No Image'
            print(Image29)
         try:
            Image30 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[30]/a/div/span/img[1]').get_attribute('src')
            print(Image30)
            len30 = 10
            letters30 = string.ascii_lowercase
            img30_name = ''.join(random.choice(letters30) for i in range(len30))
            full30_name = str(img30_name) + '.jpg'
            file30_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full30_name
            urllib.request.urlretrieve(Image30,file30_path)
         except:
            Image30 = 'No Image'
            print(Image30)
         try:
            Image31 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[31]/a/div/span/img[1]').get_attribute('src')
            print(Image31)
            len31 = 10
            letters31 = string.ascii_lowercase
            img31_name = ''.join(random.choice(letters31) for i in range(len31))
            full31_name = str(img31_name) + '.jpg'
            file31_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full31_name
            urllib.request.urlretrieve(Image31,file31_path)
         except:
            Image31 = 'No Image'
            print(Image31)
         try:
            Image32 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[32]/a/div/span/img[1]').get_attribute('src')
            print(Image32)
            len32 = 10
            letters32 = string.ascii_lowercase
            img32_name = ''.join(random.choice(letters32) for i in range(len32))
            full32_name = str(img32_name) + '.jpg'
            file32_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full32_name
            urllib.request.urlretrieve(Image32,file32_path)
         except:
            Image32 = 'No Image'
            print(Image32)
         try:
            Image33 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[33]/a/div/span/img[1]').get_attribute('src')
            print(Image33)
            len33 = 10
            letters33 = string.ascii_lowercase
            img33_name = ''.join(random.choice(letters33) for i in range(len33))
            full33_name = str(img33_name) + '.jpg'
            file33_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full33_name
            urllib.request.urlretrieve(Image33,file33_path)
         except:
            Image33 = 'No Image'
            print(Image33)
         try:
            Image34 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[34]/a/div/span/img[1]').get_attribute('src')
            print(Image34)
            len34 = 10
            letters34 = string.ascii_lowercase
            img34_name = ''.join(random.choice(letters34) for i in range(len34))
            full34_name = str(img34_name) + '.jpg'
            file34_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full34_name
            urllib.request.urlretrieve(Image34,file34_path)
         except:
            Image34 = 'No Image'
            print(Image34)
         try:
            Image35 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[35]/a/div/span/img[1]').get_attribute('src')
            print(Image35)
            len35 = 10
            letters35 = string.ascii_lowercase
            img35_name = ''.join(random.choice(letters35) for i in range(len35))
            full35_name = str(img35_name) + '.jpg'
            file35_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full35_name
            urllib.request.urlretrieve(Image35,file35_path)
         except:
            Image35 = 'No Image'
            print(Image35)
         try:
            Image36 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[36]/a/div/span/img[1]').get_attribute('src')
            print(Image36)
            len36 = 10
            letters36 = string.ascii_lowercase
            img36_name = ''.join(random.choice(letters36) for i in range(len36))
            full36_name = str(img36_name) + '.jpg'
            file36_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full36_name
            urllib.request.urlretrieve(Image36,file36_path)
         except:
            Image36 = 'No Image'
            print(Image36)
         try:
            Image37 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[37]/a/div/span/img[1]').get_attribute('src')
            print(Image37)
            len37 = 10
            letters37 = string.ascii_lowercase
            img37_name = ''.join(random.choice(letters37) for i in range(len37))
            full37_name = str(img37_name) + '.jpg'
            file37_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full37_name
            urllib.request.urlretrieve(Image37,file37_path)
         except:
            Image37 = 'No Image'
            print(Image37)
         try:
            Image38 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[38]/a/div/span/img[1]').get_attribute('src')
            print(Image38)
            len38 = 10
            letters38 = string.ascii_lowercase
            img38_name = ''.join(random.choice(letters38) for i in range(len38))
            full38_name = str(img38_name) + '.jpg'
            file38_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full38_name
            urllib.request.urlretrieve(Image38,file38_path)
         except:
            Image38 = 'No Image'
            print(Image38)									
         try:
            Image39 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[39]/a/div/span/img[1]').get_attribute('src')
            print(Image39)
            len39 = 10
            letters39 = string.ascii_lowercase
            img39_name = ''.join(random.choice(letters39) for i in range(len39))
            full39_name = str(img39_name) + '.jpg'
            file39_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full39_name
            urllib.request.urlretrieve(Image39,file39_path)
         except:
            Image39 = 'No Image'
            print(Image39)
         try:
            Image40 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[40]/a/div/span/img[1]').get_attribute('src')
            print(Image40)
            len40 = 10
            letters40 = string.ascii_lowercase
            img40_name = ''.join(random.choice(letters40) for i in range(len40))
            full40_name = str(img40_name) + '.jpg'
            file40_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full40_name
            urllib.request.urlretrieve(Image40,file40_path)
         except:
            Image40 = 'No Image'
            print(Image40)
         try:
            Image41 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[41]/a/div/span/img[1]').get_attribute('src')
            print(Image41)
            len41 = 10
            letters41 = string.ascii_lowercase
            img41_name = ''.join(random.choice(letters41) for i in range(len41))
            full41_name = str(img41_name) + '.jpg'
            file41_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full41_name
            urllib.request.urlretrieve(Image41,file41_path)
         except:
            Image41 = 'No Image'
            print(Image41)			
         try:
            Image42 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[42]/a/div/span/img[1]').get_attribute('src')
            print(Image42)
            len42 = 10
            letters42 = string.ascii_lowercase
            img42_name = ''.join(random.choice(letters42) for i in range(len42))
            full42_name = str(img42_name) + '.jpg'
            file42_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full42_name
            urllib.request.urlretrieve(Image42,file42_path)
         except:
            Image42 = 'No Image'
            print(Image42)
         try:
            Image43 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[43]/a/div/span/img[1]').get_attribute('src')
            print(Image43)
            len43 = 10
            letters43 = string.ascii_lowercase
            img43_name = ''.join(random.choice(letters43) for i in range(len43))
            full43_name = str(img43_name) + '.jpg'
            file43_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full43_name
            urllib.request.urlretrieve(Image43,file43_path)
         except:
            Image43 = 'No Image'
            print(Image43)
         try:
            Image44 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[44]/a/div/span/img[1]').get_attribute('src')
            print(Image44)
            len44 = 10
            letters44 = string.ascii_lowercase
            img44_name = ''.join(random.choice(letters44) for i in range(len44))
            full44_name = str(img44_name) + '.jpg'
            file44_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full44_name
            urllib.request.urlretrieve(Image44,file44_path)
         except:
            Image44 = 'No Image'
            print(Image44)
         try:
            Image45 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[45]/a/div/span/img[1]').get_attribute('src')
            print(Image45)
            len45 = 10
            letters45 = string.ascii_lowercase
            img45_name = ''.join(random.choice(letters45) for i in range(len45))
            full45_name = str(img45_name) + '.jpg'
            file45_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full45_name
            urllib.request.urlretrieve(Image45,file45_path)
         except:
            Image45 = 'No Image'
            print(Image45)
         try:
            Image46 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[46]/a/div/span/img[1]').get_attribute('src')
            print(Image46)
            len46 = 10
            letters46 = string.ascii_lowercase
            img46_name = ''.join(random.choice(letters46) for i in range(len46))
            full46_name = str(img46_name) + '.jpg'
            file46_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full46_name
            urllib.request.urlretrieve(Image46,file46_path)
         except:
            Image46 = 'No Image'
            print(Image46)
         try:
            Image47 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[47]/a/div/span/img[1]').get_attribute('src')
            print(Image47)
            len47 = 10
            letters47 = string.ascii_lowercase
            img47_name = ''.join(random.choice(letters47) for i in range(len47))
            full47_name = str(img47_name) + '.jpg'
            file47_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full47_name
            urllib.request.urlretrieve(Image47,file47_path)
         except:
            Image47 = 'No Image'
            print(Image47)
         try:
            Image48 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[48]/a/div/span/img[1]').get_attribute('src')
            print(Image48)
            len48 = 10
            letters48 = string.ascii_lowercase
            img48_name = ''.join(random.choice(letters48) for i in range(len48))
            full48_name = str(img48_name) + '.jpg'
            file48_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full48_name
            urllib.request.urlretrieve(Image48,file48_path)
         except:
            Image48 = 'No Image'
            print(Image48)
         try:
            Image49 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[49]/a/div/span/img[1]').get_attribute('src')
            print(Image49)
            len49 = 10
            letters49 = string.ascii_lowercase
            img49_name = ''.join(random.choice(letters49) for i in range(len49))
            full49_name = str(img49_name) + '.jpg'
            file49_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full49_name
            urllib.request.urlretrieve(Image49,file49_path)
         except:
            Image49 = 'No Image'
            print(Image49)
         try:
            Image50 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[50]/a/div/span/img[1]').get_attribute('src')
            print(Image50)
            len50 = 10
            letters50 = string.ascii_lowercase
            img50_name = ''.join(random.choice(letters50) for i in range(len50))
            full50_name = str(img50_name) + '.jpg'
            file50_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full50_name
            urllib.request.urlretrieve(Image50,file50_path)
         except:
            Image50 = 'No Image'
            print(Image50)
         try:
            Image51 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[51]/a/div/span/img[1]').get_attribute('src')
            print(Image51)
            len51 = 10
            letters51 = string.ascii_lowercase
            img51_name = ''.join(random.choice(letters51) for i in range(len51))
            full51_name = str(img51_name) + '.jpg'
            file51_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full51_name
            urllib.request.urlretrieve(Image51,file51_path)
         except:
            Image51 = 'No Image'
            print(Image51)
         try:
            Image52 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[52]/a/div/span/img[1]').get_attribute('src')
            print(Image52)
            len52 = 10
            letters52 = string.ascii_lowercase
            img52_name = ''.join(random.choice(letters52) for i in range(len52))
            full52_name = str(img52_name) + '.jpg'
            file52_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full52_name
            urllib.request.urlretrieve(Image52,file52_path)
         except:
            Image52 = 'No Image'
            print(Image52)
         try:
            Image53 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[53]/a/div/span/img[1]').get_attribute('src')
            print(Image53)
            len53 = 10
            letters53 = string.ascii_lowercase
            img53_name = ''.join(random.choice(letters53) for i in range(len53))
            full53_name = str(img53_name) + '.jpg'
            file53_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full53_name
            urllib.request.urlretrieve(Image53,file53_path)
         except:
            Image53 = 'No Image'
            print(Image53)
         try:
            Image54 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[54]/a/div/span/img[1]').get_attribute('src')
            print(Image54)
            len54 = 10
            letters54 = string.ascii_lowercase
            img54_name = ''.join(random.choice(letters54) for i in range(len54))
            full54_name = str(img54_name) + '.jpg'
            file54_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full54_name
            urllib.request.urlretrieve(Image54,file54_path)
         except:
            Image54 = 'No Image'
            print(Image54)
         try:
            Image55 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[55]/a/div/span/img[1]').get_attribute('src')
            print(Image55)
            len55 = 10
            letters55 = string.ascii_lowercase
            img55_name = ''.join(random.choice(letters55) for i in range(len55))
            full55_name = str(img55_name) + '.jpg'
            file55_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full55_name
            urllib.request.urlretrieve(Image55,file55_path)
         except:
            Image55 = 'No Image'
            print(Image55)
         try:
            Image56 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[56]/a/div/span/img[1]').get_attribute('src')
            print(Image56)
            len56 = 10
            letters56 = string.ascii_lowercase
            img56_name = ''.join(random.choice(letters56) for i in range(len56))
            full56_name = str(img56_name) + '.jpg'
            file56_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full56_name
            urllib.request.urlretrieve(Image56,file56_path)
         except:
            Image56 = 'No Image'
            print(Image56)
         try:
            Image57 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[57]/a/div/span/img[1]').get_attribute('src')
            print(Image57)
            len57 = 10
            letters57 = string.ascii_lowercase
            img57_name = ''.join(random.choice(letters57) for i in range(len57))
            full57_name = str(img57_name) + '.jpg'
            file57_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full57_name
            urllib.request.urlretrieve(Image57,file57_path)
         except:
            Image57 = 'No Image'
            print(Image57)
         try:
            Image58 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[58]/a/div/span/img[1]').get_attribute('src')
            print(Image58)
            len58 = 10
            letters58 = string.ascii_lowercase
            img58_name = ''.join(random.choice(letters58) for i in range(len58))
            full58_name = str(img58_name) + '.jpg'
            file58_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full58_name
            urllib.request.urlretrieve(Image58,file58_path)
         except:
            Image58 = 'No Image'
            print(Image58)
         try:
            Image59 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[59]/a/div/span/img[1]').get_attribute('src')
            print(Image59)
            len59 = 10
            letters59 = string.ascii_lowercase
            img59_name = ''.join(random.choice(letters59) for i in range(len59))
            full59_name = str(img59_name) + '.jpg'
            file59_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full59_name
            urllib.request.urlretrieve(Image59,file59_path)
         except:
            Image59 = 'No Image'
            print(Image59)
         try:
            Image60 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[60]/a/div/span/img[1]').get_attribute('src')
            print(Image60)
            len60 = 10
            letters60 = string.ascii_lowercase
            img60_name = ''.join(random.choice(letters60) for i in range(len60))
            full60_name = str(img60_name) + '.jpg'
            file60_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full60_name
            urllib.request.urlretrieve(Image60,file60_path)
         except:
            Image60 = 'No Image'
            print(Image60)
         try:
            Image61 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[61]/a/div/span/img[1]').get_attribute('src')
            print(Image61)
            len61 = 10
            letters61 = string.ascii_lowercase
            img61_name = ''.join(random.choice(letters61) for i in range(len61))
            full61_name = str(img61_name) + '.jpg'
            file61_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full61_name
            urllib.request.urlretrieve(Image61,file61_path)
         except:
            Image61 = 'No Image'
            print(Image61)
         try:
            Image62 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[62]/a/div/span/img[1]').get_attribute('src')
            print(Image62)
            len62 = 10
            letters62 = string.ascii_lowercase
            img62_name = ''.join(random.choice(letters62) for i in range(len62))
            full62_name = str(img62_name) + '.jpg'
            file62_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full62_name
            urllib.request.urlretrieve(Image62,file62_path)
         except:
            Image62 = 'No Image'
            print(Image62)
         try:
            Image63 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[63]/a/div/span/img[1]').get_attribute('src')
            print(Image63)
            len63 = 10
            letters63 = string.ascii_lowercase
            img63_name = ''.join(random.choice(letters63) for i in range(len63))
            full63_name = str(img63_name) + '.jpg'
            file63_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full63_name
            urllib.request.urlretrieve(Image63,file63_path)
         except:
            Image63 = 'No Image'
            print(Image63)
         try:
            Image64 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[64]/a/div/span/img[1]').get_attribute('src')
            print(Image64)
            len64 = 10
            letters64 = string.ascii_lowercase
            img64_name = ''.join(random.choice(letters64) for i in range(len64))
            full64_name = str(img64_name) + '.jpg'
            file64_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full64_name
            urllib.request.urlretrieve(Image64,file64_path)
         except:
            Image64 = 'No Image'
            print(Image64)
         try:
            Image65 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[65]/a/div/span/img[1]').get_attribute('src')
            print(Image65)
            len65 = 10
            letters65 = string.ascii_lowercase
            img65_name = ''.join(random.choice(letters65) for i in range(len65))
            full65_name = str(img65_name) + '.jpg'
            file65_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full65_name
            urllib.request.urlretrieve(Image65,file65_path)
         except:
            Image65 = 'No Image'
            print(Image65)
         try:
            Image66 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[66]/a/div/span/img[1]').get_attribute('src')
            print(Image66)
            len66 = 10
            letters66 = string.ascii_lowercase
            img66_name = ''.join(random.choice(letters66) for i in range(len66))
            full66_name = str(img66_name) + '.jpg'
            file66_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full66_name
            urllib.request.urlretrieve(Image66,file66_path)
         except:
            Image66 = 'No Image'
            print(Image66)
         try:
            Image67 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[67]/a/div/span/img[1]').get_attribute('src')
            print(Image67)
            len67 = 10
            letters67 = string.ascii_lowercase
            img67_name = ''.join(random.choice(letters67) for i in range(len67))
            full67_name = str(img67_name) + '.jpg'
            file67_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full67_name
            urllib.request.urlretrieve(Image67,file67_path)
         except:
            Image67 = 'No Image'
            print(Image67)
         try:
            Image68 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[68]/a/div/span/img[1]').get_attribute('src')
            print(Image68)
            len68 = 10
            letters68 = string.ascii_lowercase
            img68_name = ''.join(random.choice(letters68) for i in range(len68))
            full68_name = str(img68_name) + '.jpg'
            file68_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full68_name
            urllib.request.urlretrieve(Image68,file68_path)
         except:
            Image68 = 'No Image'
            print(Image68)
         try:
            Image69 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[69]/a/div/span/img[1]').get_attribute('src')
            print(Image69)
            len69 = 10
            letters69 = string.ascii_lowercase
            img69_name = ''.join(random.choice(letters69) for i in range(len69))
            full69_name = str(img69_name) + '.jpg'
            file69_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full69_name
            urllib.request.urlretrieve(Image69,file69_path)
         except:
            Image69 = 'No Image'
            print(Image69)
         try:
            Image70 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[70]/a/div/span/img[1]').get_attribute('src')
            print(Image70)
            len70 = 10
            letters70 = string.ascii_lowercase
            img70_name = ''.join(random.choice(letters70) for i in range(len70))
            full70_name = str(img70_name) + '.jpg'
            file70_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full70_name
            urllib.request.urlretrieve(Image70,file70_path)
         except:
            Image70 = 'No Image'
            print(Image70)
         try:
            Image71 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[71]/a/div/span/img[1]').get_attribute('src')
            print(Image71)
            len71 = 10
            letters71 = string.ascii_lowercase
            img71_name = ''.join(random.choice(letters71) for i in range(len71))
            full71_name = str(img71_name) + '.jpg'
            file71_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full71_name
            urllib.request.urlretrieve(Image71,file71_path)
         except:
            Image71 = 'No Image'
            print(Image71)
         try:
            Image72 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[72]/a/div/span/img[1]').get_attribute('src')
            print(Image72)
            len72 = 10
            letters72 = string.ascii_lowercase
            img72_name = ''.join(random.choice(letters72) for i in range(len72))
            full72_name = str(img72_name) + '.jpg'
            file72_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full72_name
            urllib.request.urlretrieve(Image72,file72_path)
         except:
            Image72 = 'No Image'
            print(Image72)
         try:
            Image73 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[73]/a/div/span/img[1]').get_attribute('src')
            print(Image73)
            len73 = 10
            letters73 = string.ascii_lowercase
            img73_name = ''.join(random.choice(letters73) for i in range(len73))
            full73_name = str(img73_name) + '.jpg'
            file73_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full73_name
            urllib.request.urlretrieve(Image73,file73_path)
         except:
            Image73 = 'No Image'
            print(Image73)
         try:
            Image74 = driver.find_element_by_xpath('//div[@class="flickity-slider"]/div[74]/a/div/span/img[1]').get_attribute('src')
            print(Image74)
            len74 = 10
            letters74 = string.ascii_lowercase
            img74_name = ''.join(random.choice(letters74) for i in range(len74))
            full74_name = str(img74_name) + '.jpg'
            file74_path = 'E:\\Crawling-py-Data\\extractor\\jcat\\img' + full74_name
            urllib.request.urlretrieve(Image74,file74_path)
         except:
            Image74 = 'No Image'
            print(Image74)
   
         data.append((full2_name,full3_name,full4_name,full5_name,full6_name,full7_name,full8_name,full9_name,full10_name,full12_name,full13_name,full14_name,full15_name,full16_name,full17_name,full18_name,full19_name,full20_name,full21_name,full22_name,full23_namefull24_name,full25_name,full26_name,full27_name,full28_name,full29_name,full30_name,full31_name,full32_name,full33_name,full34_name,full35_name,full36_name,full37_name,full38_namefull39_name,full40_name,full41_name,full42_name,full43_name,full44_name,full45_name,full46_name,full47_name,full48_name,full49_name,full50_name,full51_name,full52_name,full53_namefull54_name,full55_name,full56_name,full57_name,full58_name,full59_name,full60_name,full61_name,full62_name,full63_namefull64_name,full65_name,full66_name,full67_name,full68_name,full69_name,full70_name,full71_name,full72_name,full73_name,full74_name,val))
         import pandas as pd
         df = pd.DataFrame(data,columns =['Image2','Image3','Image4','Image5','Image6','Image7','Image8','Image9','Image10','Image12','Image13','Image14','Image15','Image16','Image17','Image18','Image19','Image20','Image21','Image22','Image23','Image24','Image25','Image26','Image27','Image28','Image29','Image30','Image31','Image32','Image33','Image34','Image35','Image36','Image37','Image38','Image39','Image40','Image41','Image42','Image43','Image44','Image45','Image46','Image47','Image48','Image49','Image50','Image51','Image52','Image53','Image54','Image55','Image56','Image57','Image58','Image59','Image60','Image61','Image62','Image63','Image64','Image65','Image66','Image67','Image68','Image69','Image70','Image71','Image72','Image73','Image74''PageURL'])
         df.to_csv('E:\Crawling-py-Data\extractor\jcat\jcat.csv',index=False,encoding='utf-8')
         
           
        
            
          
           
            
            
            
            

            
                        
         
        

   



