import sys, os, time
import urllib.request
import getpass
import py2exe
import requests
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from boltons import iterutils
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
wk = openpyxl.load_workbook(r'C:\Users\Bilal Shaukat\Desktop\beauti-indution.xlsx')
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column
data = []
driver = webdriver.Chrome()
driver.maximize_window()
time.sleep(2)
for i in range (1, rows+1):
         val = sh.cell(i,1).value
         driver.get(val)
         time.sleep(5)
##         catalog_search = driver.find_element_by_xpath('//div[@class="swatch-attribute-options clearfix"]/div').click()
##         catalog-search2= driver.find_element_by_xpath('//div[@class="fotorama__nav__shaft"]/div[2]/div/img')


         
         try:
            catalog_search = driver.find_element_by_xpath('//div[@class="swatch-attribute-options clearfix"]/div[1]').click()
            time.sleep(5)
            Image7 = driver.find_element_by_xpath('//div[@class="fotorama__nav__shaft"]/div[2]/div/img').get_attribute('src')
            print(Image7)
            len7 = 10
            letters7 = string.ascii_lowercase
            Image7=Image7.replace("f9c7fbe9b524c081a3ccf800cbd963eb","c687aa7517cf01e65c009f6943c2b1e9")
            img7_name = ''.join(random.choice(letters7) for i in range(len7))
            full7_name = str(img7_name) + '.jpg'
            file7_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full7_name
            urllib.request.urlretrieve(Image7,file7_path)
         except:
            Image7 = 'No Image'
            full7_name = 'No Image'
            print(Image7)
            pass
         try:
            Image8 = driver.find_element_by_xpath('//div[@class="fotorama__nav__shaft"]/div[3]/div/img').get_attribute('src')
            print(Image8)
            len8 = 10
            letters8 = string.ascii_lowercase
            Image8=Image8.replace("f9c7fbe9b524c081a3ccf800cbd963eb","c687aa7517cf01e65c009f6943c2b1e9")
            img8_name = ''.join(random.choice(letters8) for i in range(len8))
            full8_name = str(img8_name) + '.jpg'
            file8_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full8_name
            urllib.request.urlretrieve(Image8,file8_path)
         except:
            Image8 = 'No Image'
            full8_name = 'No Image'
            print(Image8)
            pass
         try:
            Image9 = driver.find_element_by_xpath('//div[@class="fotorama__nav__shaft"]/div[4]/div/img').get_attribute('src')
            print(Image9)
            len121 = 10
            letters121 = string.ascii_lowercase
            Image9=Image9.replace("f9c7fbe9b524c081a3ccf800cbd963eb","c687aa7517cf01e65c009f6943c2b1e9")
            img9_name = ''.join(random.choice(letters121) for i in range(len121))
            full9_name = str(img9_name) + '.jpg'
            file9_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full9_name
            urllib.request.urlretrieve(Image9,file9_path)
         except:
            Image9 = 'No Image'
            full9_name = 'No Image'
            print(Image9)
            pass
         time.sleep(2)
         try:
            catalog_search2 = driver.find_element_by_xpath('//div[@class="swatch-attribute-options clearfix"]/div[2]').click()
            time.sleep(5)
            Image77 = driver.find_element_by_xpath('//div[@class="fotorama__nav__shaft"]/div[2]/div/img').get_attribute('src')
            print(Image77)
            len7 = 10
            letters7 = string.ascii_lowercase
            Image77=Image77.replace("f9c7fbe9b524c081a3ccf800cbd963eb","c687aa7517cf01e65c009f6943c2b1e9")
            img7_name = ''.join(random.choice(letters7) for i in range(len7))
            full77_name = str(img7_name) + '.jpg'
            file77_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full77_name
            urllib.request.urlretrieve(Image77,file77_path)
         except:
            Image77 = 'No Image'
            full77_name = 'No Image'
            print(Image77)
            pass
         try:
            Image88 = driver.find_element_by_xpath('//div[@class="fotorama__nav__shaft"]/div[3]/div/img').get_attribute('src')
            print(Image88)
            len8 = 10
            letters8 = string.ascii_lowercase
            Image88=Image88.replace("f9c7fbe9b524c081a3ccf800cbd963eb","c687aa7517cf01e65c009f6943c2b1e9")
            img8_name = ''.join(random.choice(letters8) for i in range(len8))
            full88_name = str(img8_name) + '.jpg'
            file88_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full88_name
            urllib.request.urlretrieve(Image88,file88_path)
         except:
            Image88 = 'No Image'
            full88_name = 'No Image'
            print(Image88)
            pass
         try:
            Image99 = driver.find_element_by_xpath('//div[@class="fotorama__nav__shaft"]/div[4]/div/img').get_attribute('src')
            print(Image99)
            len121 = 10
            letters121 = string.ascii_lowercase
            Image99=Image99.replace("f9c7fbe9b524c081a3ccf800cbd963eb","c687aa7517cf01e65c009f6943c2b1e9")
            img9_name = ''.join(random.choice(letters121) for i in range(len121))
            full99_name = str(img9_name) + '.jpg'
            file99_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full99_name
            urllib.request.urlretrieve(Image99,file99_path)
         except:
            Image99 = 'No Image'
            full99_name = 'No Image'
            print(Image99)
            pass


         data.append((val,full7_name,full8_name,full9_name,full77_name,full88_name,full99_name))
         import pandas as pd
         df = pd.DataFrame(data,columns =['PageURL','Image7','Image8','Image9','Image77','Image88','Image99'])
         df.to_csv('E:\\Crawling-py-Data\\extractor\\New folder\\file.csv',index=False,encoding='utf-8')
         
