import urllib.request
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
wk = openpyxl.load_workbook(r'C:\Users\Bilal Shaukat\Desktop\beauti-indution.xlsx')
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column
data = []
driver = webdriver.Chrome()
driver.maximize_window()
#browser = webdriver.Chrome()
for i in range (1, rows+1):
         val = sh.cell(i,1).value
         time.sleep(5)
         driver.get(val)
         time.sleep(10)
         try:
            #btttn_clik = driver.find_element_by_xpath('/html/body/div[4]/div/div/div/div[1]/div/div/button').click()
            Image2 = driver.find_element_by_xpath('//div[@id="gal1"]/a[1]/img').get_attribute('src')
            print(Image2)
            len1 = 10
            letters2 = string.ascii_lowercase
            Image2=Image2.replace("-100x100.",".")
            img2_name = ''.join(random.choice(letters2) for i in range(len1)) 
            full2_name = str(img2_name) + '.jpg'
            file2_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full2_name
            urllib.request.urlretrieve(Image2,file2_path)
         except:
            Image2 = 'No Image'
            full2_name = 'No Image'
            print(Image2)
            pass
         try:
            #btttn_clik = driver.find_element_by_xpath('/html/body/div[4]/div/div/div/div[1]/div/div/button').click()
            Image3 = driver.find_element_by_xpath('//div[@id="gal1"]/a[2]/img').get_attribute('src')
            print(Image3)
            len1 = 10
            letters2 = string.ascii_lowercase
            Image3=Image3.replace("-100x100.",".")
            img3_name = ''.join(random.choice(letters2) for i in range(len1)) 
            full3_name = str(img3_name) + '.jpg'
            file3_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full3_name
            urllib.request.urlretrieve(Image3,file3_path)
         except:
            Image3 = 'No Image'
            full3_name = 'No Image'
            print(Image3)
            pass
         try:
            Image4 = driver.find_element_by_xpath('//div[@id="gal1"]/a[3]/img').get_attribute('src')
            print(Image4)
            len1 = 10
            letters2 = string.ascii_lowercase
            Image4=Image4.replace("-100x100.",".")
            img4_name = ''.join(random.choice(letters2) for i in range(len1)) 
            full4_name = str(img4_name) + '.jpg'
            file4_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full4_name
            urllib.request.urlretrieve(Image4,file4_path)
         except:
            Image4 = 'No Image'
            full4_name = 'No Image'
            print(Image4)
            pass
         try:
            Image5 = driver.find_element_by_xpath('//div[@id="gal1"]/a[4]/img').get_attribute('src')
            print(Image5)
            len1 = 10
            letters2 = string.ascii_lowercase
            Image5=Image5.replace("-100x100.",".")
            img5_name = ''.join(random.choice(letters2) for i in range(len1)) 
            full5_name = str(img5_name) + '.jpg'
            file5_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full5_name
            urllib.request.urlretrieve(Image5,file5_path)
         except:
            Image5 = 'No Image'
            full5_name = 'No Image'
            print(Image5)
            pass
         try:
            Image6 = driver.find_element_by_xpath('//div[@id="gal1"]/a[5]/img').get_attribute('src')
            print(Image6)
            len1 = 10
            letters2 = string.ascii_lowercase
            Image6=Image6.replace("-100x100.",".")
            img6_name = ''.join(random.choice(letters2) for i in range(len1)) 
            full6_name = str(img6_name) + '.jpg'
            file6_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full6_name
            urllib.request.urlretrieve(Image6,file6_path)
         except:
            Image6 = 'No Image'
            full6_name = 'No Image'
            print(Image6)
            pass
         try:
            Image7 = driver.find_element_by_xpath('//div[@id="gal1"]/a[6]/img').get_attribute('src')
            print(Image7)
            len1 = 10
            letters2 = string.ascii_lowercase
            Image7=Image7.replace("-100x100.",".")
            img7_name = ''.join(random.choice(letters2) for i in range(len1)) 
            full7_name = str(img7_name) + '.jpg'
            file7_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full7_name
            urllib.request.urlretrieve(Image7,file7_path)
         except:
            Image7 = 'No Image'
            full7_name = 'No Image'
            print(Image7)
            pass


         data.append((full2_name,full3_name,full4_name,full5_name,full6_name,full7_name,val))
import pandas as pd
df = pd.DataFrame(data,columns =['Image2','Image3','Image4','Image5','Image6','Image7','PageURL'])
df.to_csv('E:\\Crawling-py-Data\\extractor\\New folder\\file.csv',index=False,encoding='utf-8')
         
           
        
            
          
           
            
            
            
            

            
                        
         
        

   



