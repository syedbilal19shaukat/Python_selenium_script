import urllib.request
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
wk = openpyxl.load_workbook(r'C:\Users\Bilal Shaukat\Desktop\beauti-indution.xlsx')
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column
data = []
driver = webdriver.Chrome()
driver.maximize_window()
for i in range (1, rows+1):
         val = sh.cell(i,1).value
         driver.get(val)
         time.sleep(5)
         try:
             price = driver.find_element_by_xpath('//div/p/a[contains(@target, "_blank")][1]').get_attribute('href')
             print(price)
         except:
             price = 'No price'
             print(price)
             pass
         data.append((val,price))
         import pandas as pd
         df = pd.DataFrame(data,columns =['Link','Price',])
         df.to_csv('E:\\Crawling-py-Data\\extractor\\New folder\\geonline13.csv',index=False,encoding='utf-8')
         
           
        
            
          
           
            
            
            
            

            
                        
         
        

   



