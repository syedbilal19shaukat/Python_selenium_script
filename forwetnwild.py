import urllib.request
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
wk = openpyxl.load_workbook(r'C:\Users\Bilal Shaukat\Desktop\beauti-indution.xlsx')
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column
data = []
driver = webdriver.Chrome()
driver.maximize_window()
#browser = webdriver.Chrome()
for i in range (1, rows+1):
         val = sh.cell(i,1).value
         time.sleep(2)
         driver.get(val)
         cta=driver.find_element_by_xpath('//div[@class="swatch-attribute-options clearfix"]/div[1]').get_attribute('src')
##         try:
##             Image4 = driver.find_element_by_xpath('//div[@class="fotorama__nav__shaft"]/div[2]/div/img').get_attribute('src')
##             print(Image4)
##             len4 = 10
##             letters4 = string.ascii_lowercase
##             Image4=Image4.replace("jpg_80x80q80.jpg_.webp","jpg")
##             img4_name = ''.join(random.choice(letters4) for i in range(len4))
##             full4_name = str(img4_name) + '.jpg'
##             file4_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full4_name
##             urllib.request.urlretrieve(Image4,file4_path)
##         except:
##             Image4 = 'No Image'
##             full4_name = 'No Image'
##             print(Image4)
##             pass
         Image = dict()
         sampleImage = dict()
         fullname = dict()
         filepath = dict()
         try:
             j=4
             Image[j] = driver.find_element_by_xpath('//div[@class="fotorama__nav__shaft"]/div[2]/div/img').get_attribute('src')
             print(Image[j])
             len4 = 10
             letters4 = string.ascii_lowercase
             Image[j]=Image[j].replace("jpg_80x80q80.jpg_.webp","jpg")
             imgname[j] = ''.join(random.choice(letters4) for i in range(len4))
             fullname[j] = str(imgname[j]) + '.jpg'
             filepath[j] = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + fullname[j]
             urllib.request.urlretrieve(Image[j],filepath[j])
         except:
             Image[j] = 'No Image'
             fullname[j] = 'No Image'
             print(Image[j])
             pass
         try:
             j=5
             Image[j] = driver.find_element_by_xpath('//div[@class="fotorama__nav__shaft"]/div[3]/div/img').get_attribute('src')
             print(Image[j])
             len4 = 10
             letters4 = string.ascii_lowercase
             Image[j]=Image[j].replace("jpg_80x80q80.jpg_.webp","jpg")
             imgname[j] = ''.join(random.choice(letters4) for i in range(len4))
             fullname[j] = str(imgname[j]) + '.jpg'
             filepath[j] = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + fullname[j]
             urllib.request.urlretrieve(Image[j],filepath[j])
         except:
             Image[j] = 'No Image'
             fullname[j] = 'No Image'
             print(Image[j])
             pass



##         try:
##            Image5 = driver.find_element_by_xpath('//div[@class="fotorama__nav__shaft"]/div[3]/div/img').get_attribute('src')
##            print(Image5)
##            len5 = 10
##            letters5 = string.ascii_lowercase
##            Image5=Image5.replace("jpg_80x80q80.jpg_.webp","jpg")
##            img5_name = ''.join(random.choice(letters5) for i in range(len5))
##            full5_name = str(img5_name) + '.jpg'
##            file5_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full5_name
##            urllib.request.urlretrieve(Image5,file5_path)
##         except:
##            Image5 = 'No Image'
##            full5_name = 'No Image'
##            print(Image5)
##            pass
         try:
            Image6 = driver.find_element_by_xpath('//div[@class="fotorama__nav__shaft"]/div[4]/div/img').get_attribute('src')
            print(Image6)
            len6 = 10
            letters6 = string.ascii_lowercase
            Image6=Image6.replace("jpg_80x80q80.jpg_.webp","jpg")
            img6_name = ''.join(random.choice(letters6) for i in range(len6))
            full6_name = str(img6_name) + '.jpg'
            file6_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full6_name
            urllib.request.urlretrieve(Image6,file6_path)
         except:
            Image6 = 'No Image'
            full6_name = 'No Image'
            print(Image6)
            pass
##         try:
##            Image7 = driver.find_element_by_xpath('//div[@class="next-slick-list"]/div/div[7]/div/img').get_attribute('src')
##            print(Image7)
##            len7 = 10
##            letters7 = string.ascii_lowercase
##            Image7=Image7.replace("jpg_80x80q80.jpg_.webp","jpg")
##            img7_name = ''.join(random.choice(letters7) for i in range(len7))
##            full7_name = str(img7_name) + '.jpg'
##            file7_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full7_name
##            urllib.request.urlretrieve(Image7,file7_path)
##         except:
##            Image7 = 'No Image'
##            full7_name = 'No Image'
##            print(Image7)
##            pass
##         try:
##            Image8 = driver.find_element_by_xpath('//div[@class="next-slick-list"]/div/div[8]/div/img').get_attribute('src')
##            print(Image8)
##            len8 = 10
##            letters8 = string.ascii_lowercase
##            Image8=Image8.replace("jpg_80x80q80.jpg_.webp","jpg")
##            img8_name = ''.join(random.choice(letters8) for i in range(len8))
##            full8_name = str(img8_name) + '.jpg'
##            file8_path = 'E:\\Crawling-py-Data\\extractor\\New folder\\' + full8_name
##            urllib.request.urlretrieve(Image8,file8_path)
##         except:
##            Image8 = 'No Image'
##            full8_name = 'No Image'
##            print(Image8)
##            pass


         data.append((fullname[j],fullname[j],full6_name,val))
         import pandas as pd
         df = pd.DataFrame(data,columns =['Image4','Image5','Image6','PageURL'])
         df.to_csv('E:\\Crawling-py-Data\\extractor\\New folder\\file.csv',index=False,encoding='utf-8')
         
           
        
            
          
           
            
            
            
            

            
                        
         
        

   



