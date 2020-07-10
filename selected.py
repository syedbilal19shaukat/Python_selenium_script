import urllib.request
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import Select

wk = openpyxl.load_workbook(r'C:\Users\Bilal Shaukat\Desktop\beauti-indution.xlsx')
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column
data = []
driver = webdriver.Chrome()
driver.maximize_window()
for i in range (1, rows+1):
         val = sh.cell(i,1).value
         try:
             driver.get(val)
         except:
             val = 'Download Complete'
             pass
         time.sleep(12)
         element = driver.find_element_by_xpath('//*[@id="pa_size"]')
         drp=Select(element)
         time.sleep(2)
         try:
             size39 = drp.select_by_visible_text('39').text
             print(size39)
         except:
             try:
                 Size_39_stock = driver.find_element_by_xpath('//div[@class="single_variation_wrap"]/div/div/p[contains(@class,stock.out-of-stock)]').text
                 print(Size_39_stock)
             except:
                 Size_39_stock = '39 Size unavailable'
                 pass
         time.sleep(2)       
         try:
             size40 = drp.select_by_visible_text('40').text
             print(size40)
         except:
             try:
                 Size_40_stock = driver.find_element_by_xpath('//div[@class="single_variation_wrap"]/div/div/p[contains(@class,stock.out-of-stock)]').text
                 print(Size_40_stock)
             except:
                 Size_40_stock = '40 Size unavailable'
                 pass
         time.sleep(2)
         try:
             size41 = drp.select_by_visible_text('41').text
             print(size41)
         except:
             try:
                 Size_41_stock = driver.find_element_by_xpath('//div[@class="single_variation_wrap"]/div/div/p[contains(@class,stock.out-of-stock)]').text
                 print(Size_41_stock)
             except:
                 Size_41_stock = '41 Size unavailable'
                 pass
         time.sleep(2)       
         try:
             size42 = drp.select_by_visible_text('42').text
             print(size42)
         except:
             try:
                 Size_42_stock = driver.find_element_by_xpath('//div[@class="single_variation_wrap"]/div/div/p[contains(@class,stock.out-of-stock)]').text
                 print(Size_42_stock)
             except:
                 Size_42_stock = '42 Size unavailable'
                 pass
         time.sleep(2)       
         try:
             size43 = drp.select_by_visible_text('43').text
             print(size43)
         except:
             try:
                 Size_43_stock = driver.find_element_by_xpath('//div[@class="single_variation_wrap"]/div/div/p[contains(@class,stock.out-of-stock)]').text
                 print(Size_43_stock)
             except:
                 Size_43_stock = '43 Size unavailable'
                 pass
         time.sleep(2)
         try:
             size44 = drp.select_by_visible_text('44').text
             print(size44)
         except:
             try:
                 Size_44_stock = driver.find_element_by_xpath('//div[@class="single_variation_wrap"]/div/div/p[contains(@class,stock.out-of-stock)]').text
                 print(Size_44_stock)
             except:
                 Size_44_stock = '44 Size unavailable'
                 pass
         time.sleep(2)       
         try:
             size45 = drp.select_by_visible_text('45').text
             print(size45)
         except:
             try:
                 Size_45_stock = driver.find_element_by_xpath('//div[@class="single_variation_wrap"]/div/div/p[contains(@class,stock.out-of-stock)]').text
                 print(Size_45_stock)
             except:
                 Size_45_stock = '45 Size unavailable'
                 pass
         time.sleep(2)       
         try:
             size46 = drp.select_by_visible_text('46').text
             print(size46)
         except:
             try:
                 Size_46_stock = driver.find_element_by_xpath('//div[@class="single_variation_wrap"]/div/div/p[contains(@class,stock.out-of-stock)]').text
                 print(Size_46_stock)
             except:
                 Size_46_stock = '46 Size unavailable'
                 pass



         data.append((val,Size_39_stock,Size_40_stock,Size_41_stock,Size_42_stock,Size_43_stock,Size_44_stock,Size_45_stock,Size_46_stock))
         import pandas as pd
         df = pd.DataFrame(data,columns =['PageURL','Size_39_stock','Size_40_stock','Size_41_stock','Size_42_stock','Size_43_stock','Size_44_stock','Size_45_stock','Size_46_stock'])
         df.to_csv('E:\\Crawling-py-Data\\extractor\\New folder\\file_fioriri.csv',index=False,encoding='utf-8')
         
           
        
            
          
           
            
            
            
            

            
                        
         
        

   



