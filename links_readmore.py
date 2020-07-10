#----------- daraz crawl
import urllib.request
import py2exe
import openpyxl
import sys, os, time
import re
import requests
import time
from random import *
import random,string
from array import array
from bs4 import BeautifulSoup
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains

if __name__ == "__main__":
    if getattr(sys, 'frozen', False):
        chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")
        driver = webdriver.Chrome(chromedriver_path)
    else:
        driver = webdriver.Chrome()
        driver.maximize_window()
path=input("Please Enter folder path: ")   #driver.find_element_by_xpath('//div[@class="col-xs-12 text-center"]/button').click()
os.chdir(path)
filename=input("Please Enter file name: ")
pagename=input('Enter store link: ')
pages1=input('Enter starting page: ')
page1=int(pages1)
pages2=input('Enter final page: ')
page2=int(pages2)
wb=openpyxl.Workbook()
sheet=wb.active
sheet.cell(row=1, column=1).value="Links"
print("Extract store link starts...")
x=0
urls=[]
driver.maximize_window()
browser = webdriver.Chrome()
driver.maximize_window()
time.sleep(10)



for i in range(page1,page2+1):
    driver.maximize_window()    
    p=pagename.find('langFlag=en')
    if(p>1):
        newpage=pagename.replace('langFlag=en&','langFlag=en&page='+str(i)+'&')
        browser.get(newpage)
        time.sleep(3)
    else:
        browser.get(pagename+"?page="+str(i))
        time.sleep(5)

    for s in range(1,20):
        try:
            btttn_clik = browser.find_element_by_xpath('//div[@class="col-xs-12 text-center"]/button').click()
            time.sleep(15)
            mypd = browser.find_element_by_tag_name('html')
            mypd.send_keys(Keys.DOWN)  
            time.sleep(5)
        except:
            pass

    for j in range(1,100):
        k=str(j)
        d='//div[@class="products-list"]/div['
        e=']/div[@class="product-preview"]/h3/a'
        f=d+k+e
        try:
            dpage=browser.find_element_by_xpath(f).get_attribute('href')
        except:
            pass
        urls.insert(0,str(dpage))
urls=list(set(urls))
#browser.quit()
for i in range(len(urls)):
    print("Link: "+str(i+1))
    sheet.cell(row=x+2, column=1).value=urls[i]
    x=x+1
    wb.save(filename+'.xlsx')
wb.save(filename+'.xlsx')
print("Download Compeltes...")
