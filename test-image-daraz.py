#----------- daraz crawl
import openpyxl
import sys, os, time
import re
import requests
from random import *
import random,string
from array import array
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
if __name__ == "__main__":
    if getattr(sys, 'frozen', False):
        chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")
        driver = webdriver.Chrome(chromedriver_path)
    else:
        driver = webdriver.Chrome()
path=input("Please Enter folder path: ")
os.chdir(path)
filename=input("Please Enter file name: ")
pagename=input('Enter store link: ')
pages1=input('Enter starting page: ')
page1=int(pages1)
pages2=input('Enter final page: ')
page2=int(pages2)
wb=openpyxl.Workbook()
sheet=wb.active
sheet.cell(row=1, column=1).value="Links"
sheet.cell(row=1, column=2).value="name"
sheet.cell(row=1, column=3).value="brand"
sheet.cell(row=1, column=4).value="primary_category"
sheet.cell(row=1, column=5).value="promotion_start_date"
sheet.cell(row=1, column=6).value="promotion_price"
sheet.cell(row=1, column=7).value="promotion_msrp"
sheet.cell(row=1, column=8).value="promotion_quantity"
sheet.cell(row=1, column=9).value="promotion_end_date"
sheet.cell(row=1, column=10).value="promotion_payment_method"
sheet.cell(row=1, column=11).value="promotion_uses_per_customer"
sheet.cell(row=1, column=12).value="promotion_type"
sheet.cell(row=1, column=13).value="promotion_bins"
sheet.cell(row=1, column=14).value="variation"
sheet.cell(row=1, column=15).value="parent_sku"
sheet.cell(row=1, column=16).value="vendor_sku"
sheet.cell(row=1, column=17).value="quantity"
sheet.cell(row=1, column=18).value="warehouse_quantity"
sheet.cell(row=1, column=19).value="long_description"
sheet.cell(row=1, column=20).value="short_description"
sheet.cell(row=1, column=21).value="specifications"
sheet.cell(row=1, column=22).value="refund_policy"
sheet.cell(row=1, column=23).value="product_warranty_short"
sheet.cell(row=1, column=24).value="product_warranty_long"
sheet.cell(row=1, column=25).value="price"
sheet.cell(row=1, column=26).value="code"
sheet.cell(row=1, column=27).value="main_image"
sheet.cell(row=1, column=28).value="image_2"
sheet.cell(row=1, column=29).value="image_3"
sheet.cell(row=1, column=30).value="image_4"
sheet.cell(row=1, column=31).value="image_5"
sheet.cell(row=1, column=32).value="image_6"
sheet.cell(row=1, column=33).value="image_7"
sheet.cell(row=1, column=34).value="image_8"
sheet.cell(row=1, column=35).value="meta_title"
sheet.cell(row=1, column=36).value="meta_keywords"
sheet.cell(row=1, column=37).value="meta_description"
sheet.cell(row=1, column=38).value="msrp"
sheet.cell(row=1, column=39).value="search_weight"
sheet.cell(row=1, column=40).value="shipping_cost_class"
sheet.cell(row=1, column=41).value="free_shipping"
sheet.cell(row=1, column=42).value="color"
sheet.cell(row=1, column=43).value="product_weight"
sheet.cell(row=1, column=44).value="color_family"
sheet.cell(row=1, column=45).value="product_keywords"
sheet.cell(row=1, column=46).value="delivery_time"
sheet.cell(row=1, column=47).value="youtube_id"
sheet.cell(row=1, column=48).value="shipping_height"
sheet.cell(row=1, column=49).value="shipping_length"
sheet.cell(row=1, column=50).value="shipping_width"
sheet.cell(row=1, column=51).value="shipping_weight"
sheet.cell(row=1, column=52).value="sku_city"
sheet.cell(row=1, column=53).value="upc"
sheet.cell(row=1, column=54).value="ean"
sheet.cell(row=1, column=55).value="what_is_in_the_box"
sheet.cell(row=1, column=56).value="breadcrumb"
print("Extract daraz store link starts...")
x=0
urls=[]
browser = webdriver.Chrome()
for i in range(page1,page2+1):
    p=pagename.find('langFlag=en')
    if(p>1):
        newpage=pagename.replace('langFlag=en&','langFlag=en&page='+str(i)+'&')
        browser.get(newpage)
        time.sleep(3)
    else:
        browser.get(pagename+"?page="+str(i))
        time.sleep(3)
    for j in range(1,41):
        k=str(j)
        a='//*[@id="root"]/div/div[3]/div/div/div[1]/div[3]/div['
        b=']/div/div/div[1]/div/a'
        c=a+k+b
        d='//*[@id="root"]/div/div[3]/div/div/div[1]/div[2]/div['
        e=']/div/div/div[1]/div/a'
        f=d+k+e
        try:
            dpage=browser.find_element_by_path(c).get_attribute('href')
        except:
            try:
                dpage=browser.find_element_by_xpath(f).get_attribute('href')
            except:
                pass
        urls.insert(0,str(dpage))
urls=list(set(urls))
browser.quit()
for i in range(len(urls)):
    print("Link: "+str(i+1))
    sheet.cell(row=x+2, column=1).value=urls[i]
    res = requests.get(urls[i])
    time.sleep(5)
    soup = BeautifulSoup(res.content,'html.parser')
    try:
        sheet.cell(row=x+2, column=25).value=soup.find("span" , {"class" : "pdp-price pdp-price_type_deleted pdp-price_color_lightgray pdp-price_size_xs"}).text
    except:
        pass
    try:
        sheet.cell(row=x+2, column=42).value=soup.find("span" , {"class" : "sku-name "}).text
    except:
        pass
    try:
        sheet.cell(row=x+2, column=2).value=soup.find("h1" , {"class" : "pdp-product-title"}).text
    except:
        pass
    try:
        short_description=soup.find("div" , {"class" : "html-content pdp-product-highlights"})
        short_description=str(short_description)
        short_description=short_description.replace('<div class="html-content pdp-product-highlights"><ul class="">',"")
        short_description=short_description.replace('<li class="">',"<li>")
        short_description=short_description.replace('</ul></div>',"")
        sheet.cell(row=x+2, column=20).value=short_description
    except:
        pass
    try:
        description_text=soup.find("div" , {"class" : "html-content detail-content"}).text
        sheet.cell(row=x+2, column=19).value=str(description_text)
    except:
        pass
    try:
        sheet.cell(row=x+2, column=3).value=soup.find("a" , {"class" : "pdp-link pdp-link_size_s pdp-link_theme_blue pdp-product-brand__brand-link"}).text
    except:
        pass
    try:
        sheet.cell(row=x+2, column=6).value=soup.find("span" , {"class" : "pdp-price pdp-price_type_normal pdp-price_color_orange pdp-price_size_xl"}).text
    except:
        pass
    """try:
        sheet.cell(row=x+2, column=31).value=soup.find("a" , {"class" : "pdp-link pdp-link_size_l pdp-link_theme_black seller-name__detail-name"}).text
        seller=soup.find("a" , {"class" : "pdp-link pdp-link_size_l pdp-link_theme_black seller-name__detail-name"}).text
    except:
        pass"""
    try:
        warranty1 = soup.find_all("div" , {"class" : "warranty__option-item"})
        sheet.cell(row=x+2, column=23).value=warranty1[2].text
    except:
        pass
    try:
        sku=soup.find_all("div" , {"class" : "html-content key-value"})
        sheet.cell(row=x+2, column=15).value=sku[1].text
    except:
        pass
    try:
        images=soup.find_all("img" , {"class" : "pdp-mod-common-image item-gallery__thumbnail-image"})
    except:
        pass
    try:
        breadcrumb=soup.find("ul" , {"class" : "breadcrumb"}).text
        breadcrumb=breadcrumb.replace("\n","|")
        sheet.cell(row=x+2, column=56).value=str(breadcrumb)
    except:
        pass
    try:
        category=soup.find_all("li" , {"class" : "breadcrumb_item"})
        sheet.cell(row=x+2, column=4).value=str(category[len(category)-2].text)
    except:
        pass
    sizearray = []
    try:
        size = soup.find_all('span' , { 'class' : 'sku-variable-size'})
        size1 = soup.find('span' , { 'class' : 'sku-variable-size-selected'}).text
        for sizes in size:
            size2 = sizes['title']
            sizearray.append(size2)
            list(sizearray)
        sizearray.append(size1)
        Sizecol = str(sizearray)
        Sizecol = Sizecol.replace("', '"," | ")
        Sizecol = Sizecol.replace("['","")
        Sizecol = Sizecol.replace("']","")
        z=x+1
        for i in range(len(sizearray)+1):
            sheet.cell(row=z+2, column=16).value=str(sizearray[i])
            z=z+1
    except:
        if len(sizearray)==0:            
            try:
                size = soup.find_all('span' , { 'class' : 'sku-variable-name'})
                size1 = soup.find('span' , { 'class' : 'sku-variable-name-selected'}).text
                for sizes in size:
                    size2 = sizes['title']
                    sizearray.append(size2)
                    list(sizearray)
                sizearray.append(size1)
                Sizecol = str(sizearray)
                Sizecol = Sizecol.replace("', '"," | ")
                Sizecol = Sizecol.replace("['","")
                Sizecol = Sizecol.replace("']","")
                z=x+1
                for i in range(len(sizearray)+1):
                    sheet.cell(row=z+2, column=16).value=str(sizearray[i])
                    z=z+1
            except:
                pass
    try:
        whats_in_the_box=soup.find("div" , {"class" : "box-content"}).text
        whats_in_the_box=whats_in_the_box.replace("What’s in the box","")
        sheet.cell(row=x+2, column=55).value=str(whats_in_the_box)
    except:
        pass
    y=0
    for image in images:
        image_final1=image['src']
        image_final1=image_final1.replace("_120x120q75.jpg","")
        rest1=requests.get(image_final1)
        rest1.raise_for_status()
        characters = 10
        letters = string.ascii_lowercase
        img_str = ''.join(random.choice(letters) for i in range(characters))
        imageFile11 = open(str(img_str) + '.jpg','wb')
        sheet.cell(row=x+2, column=27+y).value=str(img_str) + '.jpg'
        for chunk in rest1.iter_content(10000000):
            imageFile11.write(chunk)
        y=y+1
    if len(sizearray)!=0:
        x=z
    else:
        x=x+1
    wb.save(filename+'.xlsx')
wb.save(filename+'.xlsx')
print("Download Compeltes...")
