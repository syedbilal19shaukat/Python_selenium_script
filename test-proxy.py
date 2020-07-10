import urllib.request
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
wk = openpyxl.load_workbook(r'E:\test\Bookplates.xlsx')
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column
data = []
PROXY = "72.204.208.157:8080" # IP:PORT or HOST:PORT
driver = webdriver.Chrome(executable_path=r"C:\Users\Bilal Shaukat\Downloads\chromedriver_win32\chromedriver.exe")
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--proxy-server=%s' % PROXY)
driver = webdriver.Chrome(options=chrome_options)
driver.maximize_window()
for i in range (1, rows+1):
         val = sh.cell(i,1).value
         driver.get(val)
         time.sleep(15)
         Name = driver.find_element_by_xpath('//div/h1').text
         print(Name)
         Breadcrumb = driver.find_element_by_xpath('//div[@class="breadscrumb-inner"]/ol').text
         print(Breadcrumb)
         try:
             SKU = driver.find_element_by_xpath('//div[@class="product_meta"]/span[@class="sku_wrapper"]/span').text
             print(SKU)
         except:
             SKU = 'No SKU'
             print(SKU)
         Price = driver.find_element_by_xpath('//p[@class="price"] | //p[text()="-"]/span[@class="woocommerce-Price-amount amount"]').text
         print(Price)
         try:
             Sizes = driver.find_element_by_xpath('//*[@id="tab-description"]/table/tbody/tr[2]/td[2]').text
             print(Sizes)
         except:
             Sizes = 'No Size'
             print(Sizes)
         try:   
             Colors = driver.find_element_by_xpath('//*[@id="tab-description"]/table/tbody/tr[5]/td[2]/span').text
             print(Colors)
         except:
             Colors = 'No Colors'
             print(Colors)
         try:   
             Category = driver.find_element_by_xpath('//*[@id="product-3501"]/div/div[2]/div/div[5]/span[2]/a[1]').text
             print(Category)
         except:
             Category = 'No Category'
             print(Category)
         try:
             Description = driver.find_element_by_xpath('//div[@id="tab-description"]/p[1]').text
             print(Description)
         except:
             Description = 'No Description'
             print(Description)
         try:
             Short_Description = driver.find_element_by_xpath('//div[@class="woocommerce-product-details__short-description"]/ul').text
             print(Short_Description)
         except:
             Short_Description = 'No short Description'
             print(Short_Description)             
         try:
            Image1 = driver.find_element_by_xpath('//figure[@class="woocommerce-product-gallery__wrapper"]/div[1]/img').get_attribute('src')
            print(Image1)
            len1 = 10
            letters1 = string.ascii_lowercase
            img1_name = ''.join(random.choice(letters1) for i in range(len1))
            full1_name = str(img1_name) + '.jpg'
            file1_path = 'E:\\test\\final\\img\\' + full1_name
            urllib.request.urlretrieve(Image1,file1_path)
         except:
            Image1 = 'No Image'
            print(Image1)
         try:
            Image2 = driver.find_element_by_xpath('//figure[@class="woocommerce-product-gallery__wrapper"]/div[2]/img').get_attribute('src')
            print(Image2)
            len2 = 10
            letters2 = string.ascii_lowercase
            img2_name = ''.join(random.choice(letters2) for i in range(len2))
            full2_name = str(img2_name) + '.jpg'
            file2_path = 'E:\\test\\final\\img\\' + full2_name
            urllib.request.urlretrieve(Image2,file2_path)
         except:
            Image2 = 'No Image'
            print(Image2)
         try:
            Image3 = driver.find_element_by_xpath('//figure[@class="woocommerce-product-gallery__wrapper"]/div[3]/img').get_attribute('src')
            print(Image3)
            len3 = 10
            letters3 = string.ascii_lowercase
            img3_name = ''.join(random.choice(letters3) for i in range(len3))
            full3_name = str(img3_name) + '.jpg'
            file3_path = 'E:\\test\\final\\img\\' + full3_name
            urllib.request.urlretrieve(Image3,file3_path)
         except:
            Image3 = 'No Image'
            print(Image3)
         try:
            Image4 = driver.find_element_by_xpath('//figure[@class="woocommerce-product-gallery__wrapper"]/div[4]/img').get_attribute('src')
            print(Image4)
            len4 = 10
            letters4 = string.ascii_lowercase
            img4_name = ''.join(random.choice(letters4) for i in range(len4))
            full4_name = str(img4_name) + '.jpg'
            file4_path = 'E:\\test\\final\\img\\' + full4_name
            urllib.request.urlretrieve(Image4,file4_path)
         except:
            Image4 = 'No Image'
            print(Image4)

         data.append((Name,Breadcrumb,SKU,Sizes,Colors,Category,Short_Description,Description,Image1,Image2,Image3,Image4,val))
         import pandas as pd
         df = pd.DataFrame(data,columns =['Name','Breadcrumb','SKU','Sizes','Colors','Category','Short_Description','Description','Image1','Image2','Image3','Image4','PageURL'])
         df.to_csv('E:\\test\\final\\test-data.csv',index=False,encoding='utf-8')
         
           
        
            
          
           
            
            
            
            

            
                        
         
        

   



