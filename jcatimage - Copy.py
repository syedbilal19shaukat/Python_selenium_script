import urllib.request
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
wk = openpyxl.load_workbook(r'C:\Users\Bilal Shaukat\Desktop\beauti-indution.xlsx')
sh = wk['Sheet1']
rows = sh.max_row
col = sh.max_column
data = []
driver = webdriver.Chrome()
driver.maximize_window()
for i in range (1, rows+1):
         val = sh.cell(i,1).value
         driver.get(val)
         time.sleep(10)
         try:
            Image2 = driver.find_element_by_xpath('//div[@class="fotorama__stage__shaft fotorama__grab"]/div[1]/img').get_attribute('src')
            print(Image2)
            len2 = 10
            letters2 = string.ascii_lowercase
            img2_name = ''.join(random.choice(letters2) for i in range(len2))
            full2_name = str(img2_name) + '.jpg'
            file2_path = '\\Crawling-py-Data\\extractor\\New folder\\' + full2_name
            urllib.request.urlretrieve(Image2,file2_path)
         except:
            Image2 = 'No Image'
            print(Image2)
         try:
            Image3 = driver.find_element_by_xpath('//div[@class="fotorama__stage__shaft fotorama__grab"]/div[2]/img').get_attribute('src')
            print(Image3)
            len2 = 10
            letters3 = string.ascii_lowercase
            img3_name = ''.join(random.choice(letters3) for i in range(len2))
            full3_name = str(img3_name) + '.jpg'
            file3_path = '\\Crawling-py-Data\\extractor\\New folder\\' + full3_name
            urllib.request.urlretrieve(Image3,file3_path)
         except:
            Image3 = 'No Image'
            print(Image3)
         try:
            Image4 = driver.find_element_by_xpath('//div[@class="fotorama__stage__shaft fotorama__grab"]/div[3]/img').get_attribute('src')
            print(Image4)
            len2 = 10
            letters4 = string.ascii_lowercase
            img4_name = ''.join(random.choice(letters4) for i in range(len2))
            full4_name = str(img4_name) + '.jpg'
            file4_path = '\\Crawling-py-Data\\extractor\\New folder\\' + full4_name
            urllib.request.urlretrieve(Image4,file4_path)
         except:
            Image4 = 'No Image'
            print(Image4)
         try:
             Description = driver.find_element_by_xpath('//div[@class="product attribute description"]/div[@class="value"]/p').text
             print(Description)
         except:
             Description = 'No Description'
             print(Description)
             pass
         try:
             specs = driver.find_element_by_xpath('//div[@class="product attribute claims"]/div').text
             print(specs)
         except:
             specs = 'No Description'
             print(specs)
             pass


         data.append((full2_name,full3_name,full4_name,Description,specs,val))
         import pandas as pd
         df = pd.DataFrame(data,columns =['Image2','Image3','Image4','Description','specs','PageURL'])
         df.to_csv('E:\\Crawling-py-Data\\extractor\\New folder\\file.csv',index=False,encoding='utf-8')
         
           
        
            
          
           
            
            
            
            

            
                        
         
        

   



